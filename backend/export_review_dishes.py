#!/usr/bin/env python3
"""Export dishes mentioned in the top 5 Google reviews for each restaurant.

This script:
1. Reads restaurant records from the Supabase `places` table.
2. Pulls up to 5 Google Places reviews per restaurant.
3. Extracts likely dish names from those reviews using rule-based parsing.
4. Writes the results to an Excel workbook with summary + review detail sheets.

Run:
    python export_review_dishes.py
    python export_review_dishes.py --limit 25
    python export_review_dishes.py --output ../datasets/review_dishes.xlsx

Required environment variables in backend/.env:
    SUPABASE_URL=...
    SUPABASE_ANON_KEY=...
    GOOGLE_PLACES_API_KEY=...
"""

from __future__ import annotations

import argparse
import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Iterable

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import create_client


load_dotenv(Path(__file__).with_name(".env"))

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_ANON_KEY"]
GOOGLE_API_KEY = os.environ["GOOGLE_PLACES_API_KEY"]

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

PLACES_NEW_BASE = "https://places.googleapis.com/v1/places"
PLACES_FIELD_MASK = ",".join(
    [
        "displayName",
        "formattedAddress",
        "rating",
        "userRatingCount",
        "googleMapsUri",
        "reviews",
    ]
)

TOKEN_RE = re.compile(r"[a-zA-Z][a-zA-Z&'/.-]*")
REVIEW_SPLIT_RE = re.compile(r"[.!?;\n]+")

CONNECTOR_WORDS = {
    "and",
    "or",
    "with",
    "plus",
}

LEADING_TRIM_WORDS = {
    "a",
    "an",
    "the",
    "their",
    "our",
    "my",
    "your",
    "this",
    "that",
    "these",
    "those",
    "some",
    "signature",
    "famous",
    "popular",
    "crispy",
    "spicy",
    "delicious",
    "amazing",
    "good",
    "great",
    "best",
    "must",
}

TRAILING_TRIM_WORDS = {
    "here",
    "there",
    "though",
    "too",
    "today",
    "again",
}

BAD_DISH_PHRASES = {
    "customer service",
    "portion size",
    "waiting time",
    "long queue",
    "ice water",
    "dining area",
    "friendly staff",
    "service staff",
    "coffee shop",
    "food court",
    "nice place",
    "great place",
    "good food",
    "great food",
}

FOOD_TERMS = {
    "bagel",
    "bak",
    "bao",
    "bbq",
    "beef",
    "bento",
    "biryani",
    "biscuit",
    "broth",
    "burger",
    "burrito",
    "cake",
    "carbonara",
    "carrot",
    "cereal",
    "char",
    "cheese",
    "chicken",
    "chips",
    "chilli",
    "clam",
    "claypot",
    "coffee",
    "congee",
    "cookie",
    "crab",
    "croissant",
    "curry",
    "dessert",
    "dim",
    "dish",
    "don",
    "donburi",
    "dumpling",
    "egg",
    "fish",
    "fries",
    "gelato",
    "goza",
    "gyoza",
    "hor",
    "hotpot",
    "ice",
    "katsu",
    "kaya",
    "laksa",
    "latte",
    "lobster",
    "mee",
    "mie",
    "milkshake",
    "murtabak",
    "naan",
    "nasi",
    "noodle",
    "noodles",
    "omelette",
    "omurice",
    "oyster",
    "pancake",
    "pasta",
    "pho",
    "pizza",
    "porridge",
    "prata",
    "prawn",
    "ramen",
    "rendang",
    "rice",
    "risotto",
    "rojak",
    "roti",
    "salad",
    "sandwich",
    "satay",
    "sausage",
    "scallop",
    "set",
    "shakshuka",
    "siew",
    "soup",
    "steak",
    "sundae",
    "sushi",
    "taco",
    "takoyaki",
    "teh",
    "tempura",
    "thosai",
    "toast",
    "tofu",
    "udon",
    "waffle",
    "wanton",
    "wings",
    "wrap",
    "yakitori",
    "yoghurt",
    "yogurt",
}

BLOCKED_CANDIDATE_WORDS = {
    "again",
    "also",
    "are",
    "as",
    "at",
    "be",
    "because",
    "been",
    "but",
    "for",
    "from",
    "here",
    "if",
    "into",
    "is",
    "it",
    "its",
    "just",
    "of",
    "on",
    "really",
    "so",
    "super",
    "than",
    "that",
    "theirs",
    "them",
    "they",
    "this",
    "those",
    "too",
    "very",
    "was",
    "were",
    "which",
    "while",
    "with",
}

KNOWN_DISHES = {
    "bak chor mee": "Bak Chor Mee",
    "bak kut teh": "Bak Kut Teh",
    "banana leaf rice": "Banana Leaf Rice",
    "banh mi": "Banh Mi",
    "beef hor fun": "Beef Hor Fun",
    "beef noodle soup": "Beef Noodle Soup",
    "beef noodles": "Beef Noodles",
    "beef pho": "Beef Pho",
    "beef rendang": "Beef Rendang",
    "beef rice bowl": "Beef Rice Bowl",
    "beef steak": "Beef Steak",
    "beef burger": "Beef Burger",
    "bibimbap": "Bibimbap",
    "biryani": "Biryani",
    "breakfast platter": "Breakfast Platter",
    "butter chicken": "Butter Chicken",
    "carbonara": "Carbonara",
    "carrot cake": "Carrot Cake",
    "char kway teow": "Char Kway Teow",
    "char siew rice": "Char Siew Rice",
    "cheese fries": "Cheese Fries",
    "chicken biryani": "Chicken Biryani",
    "chicken chop": "Chicken Chop",
    "chicken katsu": "Chicken Katsu",
    "chicken rice": "Chicken Rice",
    "chicken satay": "Chicken Satay",
    "chicken wings": "Chicken Wings",
    "chilli crab": "Chilli Crab",
    "crab pasta": "Crab Pasta",
    "croissant": "Croissant",
    "curry chicken": "Curry Chicken",
    "curry laksa": "Curry Laksa",
    "dim sum": "Dim Sum",
    "double cheeseburger": "Double Cheeseburger",
    "dry ban mian": "Dry Ban Mian",
    "dumplings": "Dumplings",
    "egg tart": "Egg Tart",
    "eggs benedict": "Eggs Benedict",
    "fish and chips": "Fish And Chips",
    "fish soup": "Fish Soup",
    "fried chicken": "Fried Chicken",
    "fried rice": "Fried Rice",
    "fried tofu": "Fried Tofu",
    "garlic naan": "Garlic Naan",
    "gelato": "Gelato",
    "gyoza": "Gyoza",
    "hainanese chicken rice": "Hainanese Chicken Rice",
    "hokkien mee": "Hokkien Mee",
    "hot and sour soup": "Hot And Sour Soup",
    "hotpot": "Hotpot",
    "ice cream": "Ice Cream",
    "katsu curry": "Katsu Curry",
    "kaya toast": "Kaya Toast",
    "laksa": "Laksa",
    "lasagna": "Lasagna",
    "latte": "Latte",
    "mala xiang guo": "Mala Xiang Guo",
    "margherita pizza": "Margherita Pizza",
    "mee goreng": "Mee Goreng",
    "mee rebus": "Mee Rebus",
    "miso ramen": "Miso Ramen",
    "murtabak": "Murtabak",
    "naan": "Naan",
    "nasi goreng": "Nasi Goreng",
    "nasi lemak": "Nasi Lemak",
    "omelette rice": "Omelette Rice",
    "omurice": "Omurice",
    "oyster omelette": "Oyster Omelette",
    "pad thai": "Pad Thai",
    "pancakes": "Pancakes",
    "pasta": "Pasta",
    "pepperoni pizza": "Pepperoni Pizza",
    "pho": "Pho",
    "pizza": "Pizza",
    "poke bowl": "Poke Bowl",
    "pork belly": "Pork Belly",
    "pork chop": "Pork Chop",
    "prawn noodles": "Prawn Noodles",
    "prata": "Prata",
    "quesadilla": "Quesadilla",
    "ramen": "Ramen",
    "rendang": "Rendang",
    "risotto": "Risotto",
    "roasted duck": "Roasted Duck",
    "rojak": "Rojak",
    "roti prata": "Roti Prata",
    "salmon sashimi": "Salmon Sashimi",
    "salted egg chicken": "Salted Egg Chicken",
    "satay": "Satay",
    "scallop pasta": "Scallop Pasta",
    "seafood pasta": "Seafood Pasta",
    "shakshuka": "Shakshuka",
    "soup dumplings": "Soup Dumplings",
    "steak": "Steak",
    "sukiyaki": "Sukiyaki",
    "sushi": "Sushi",
    "tacos": "Tacos",
    "takoyaki": "Takoyaki",
    "tempura": "Tempura",
    "thosai": "Thosai",
    "tom yum soup": "Tom Yum Soup",
    "truffle fries": "Truffle Fries",
    "udon": "Udon",
    "waffles": "Waffles",
    "wanton mee": "Wanton Mee",
    "xiao long bao": "Xiao Long Bao",
    "yakitori": "Yakitori",
}

KNOWN_DISH_PATTERNS = [
    (re.compile(rf"\b{re.escape(dish)}\b"), display)
    for dish, display in sorted(KNOWN_DISHES.items(), key=lambda item: len(item[0]), reverse=True)
]

PATTERN_PREFIXES = [
    r"(?:ordered|order|got|get|had|have|tried|try|ate|eat|enjoyed|enjoy|loved|love|liked|like)",
    r"(?:recommend|recommended|recommending)",
    r"(?:must\s+try|don't\s+miss|do\s+try|go\s+for)",
    r"(?:favourite|favorite|best|signature|popular)",
]
PATTERN_RE = re.compile(
    rf"\b(?:{'|'.join(PATTERN_PREFIXES)})\b\s+(?:the\s+)?([a-zA-Z][a-zA-Z&'/.-]*(?:\s+[a-zA-Z][a-zA-Z&'/.-]*){{0,4}})",
    re.IGNORECASE,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pull top Google reviews for restaurants and export extracted dishes to Excel."
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Number of restaurants to process. If omitted, you will be prompted in the terminal.",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output Excel path. Defaults to datasets/review_dishes_<timestamp>.xlsx",
    )
    return parser.parse_args()


def prompt_for_limit(default: int = 10) -> int | None:
    while True:
        raw = input(
            f"How many restaurants do you want to extract from? "
            f"(enter a number, or 0 for all) [{default}]: "
        ).strip()
        if not raw:
            return default
        try:
            value = int(raw)
        except ValueError:
            print("Please enter a whole number, for example 10 or 25.")
            continue
        if value < 0:
            print("Please enter 0 or a positive number.")
            continue
        return None if value == 0 else value


def resolve_output_path(output_arg: str | None) -> Path:
    if output_arg:
        return Path(output_arg).expanduser().resolve()
    project_root = Path(__file__).resolve().parents[1]
    datasets_dir = project_root / "datasets"
    datasets_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return datasets_dir / f"review_dishes_{stamp}.xlsx"


def fetch_restaurants(limit: int | None) -> list[dict]:
    query = supabase.table("places").select(
        "id,name,address,gmaps_place_id,rating,user_rating_count"
    ).order("id", desc=False)
    if limit is not None:
        query = query.limit(limit)
    rows = query.execute().data or []
    return [row for row in rows if row.get("gmaps_place_id")]


def fetch_place_reviews(place_id: str) -> dict:
    url = f"{PLACES_NEW_BASE}/{place_id}"
    headers = {
        "X-Goog-Api-Key": GOOGLE_API_KEY,
        "X-Goog-FieldMask": PLACES_FIELD_MASK,
    }
    response = requests.get(url, headers=headers, timeout=20)
    response.raise_for_status()
    return response.json()


def tokens(text: str) -> list[str]:
    return TOKEN_RE.findall(text.lower())


def pretty_label(text: str) -> str:
    return " ".join(word.capitalize() for word in text.split())


def normalize_candidate(candidate: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9&'/\-\s]", " ", candidate.lower())
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -")
    parts = cleaned.split()
    while parts and parts[0] in LEADING_TRIM_WORDS:
        parts.pop(0)
    while parts and parts[-1] in TRAILING_TRIM_WORDS:
        parts.pop()
    if not parts:
        return ""
    if len(parts) > 4:
        return ""
    if all(part in CONNECTOR_WORDS for part in parts):
        return ""
    if any(part in BLOCKED_CANDIDATE_WORDS for part in parts):
        return ""
    normalized = " ".join(parts)
    if normalized in BAD_DISH_PHRASES:
        return ""
    if not any(part in FOOD_TERMS for part in parts):
        return ""
    return normalized


def split_candidate(candidate: str) -> Iterable[str]:
    for part in re.split(r",|/|\band\b|\bor\b", candidate, flags=re.IGNORECASE):
        normalized = normalize_candidate(part)
        if normalized:
            yield normalized


def extract_known_dishes(text: str) -> set[str]:
    lowered = text.lower()
    found: set[str] = set()
    for pattern, display in KNOWN_DISH_PATTERNS:
        if pattern.search(lowered):
            found.add(display)
    return found


def extract_pattern_dishes(text: str) -> set[str]:
    found: set[str] = set()
    for sentence in REVIEW_SPLIT_RE.split(text):
        sentence = sentence.strip()
        if not sentence:
            continue
        for match in PATTERN_RE.finditer(sentence):
            for candidate in split_candidate(match.group(1)):
                found.add(KNOWN_DISHES.get(candidate, pretty_label(candidate)))
    return found


def extract_dishes(review_text: str) -> list[str]:
    dishes = extract_known_dishes(review_text)
    dishes.update(extract_pattern_dishes(review_text))
    prioritized = sorted(dishes, key=lambda item: (-len(item.split()), item))
    kept: list[str] = []
    kept_norms: list[str] = []
    for dish in prioritized:
        norm = dish.lower()
        if any(re.search(rf"\b{re.escape(norm)}\b", existing) for existing in kept_norms):
            continue
        kept.append(dish)
        kept_norms.append(norm)
    return sorted(kept)


def flatten_review_text(review: dict) -> str:
    text = review.get("text")
    if isinstance(text, dict):
        return str(text.get("text") or "").strip()
    return str(text or "").strip()


def build_restaurant_summary(restaurant: dict, details: dict) -> tuple[dict, list[dict]]:
    reviews = details.get("reviews") or []
    dish_counter: Counter[str] = Counter()
    review_rows: list[dict] = []

    for index, review in enumerate(reviews[:5], start=1):
        text = flatten_review_text(review)
        dishes = extract_dishes(text)
        for dish in dishes:
            dish_counter[dish] += 1

        review_rows.append(
            {
                "Restaurant ID": restaurant.get("id"),
                "Restaurant Name": restaurant.get("name"),
                "Review Rank": index,
                "Review Rating": review.get("rating"),
                "Review Publish Time": review.get("publishTime"),
                "Review Author": (review.get("authorAttribution") or {}).get("displayName"),
                "Extracted Dishes": ", ".join(dishes) if dishes else "",
                "Review Text": text,
            }
        )

    top_dishes = ", ".join(
        f"{dish} ({count})" for dish, count in dish_counter.most_common(10)
    )
    summary_row = {
        "Restaurant ID": restaurant.get("id"),
        "Restaurant Name": restaurant.get("name"),
        "Address": details.get("formattedAddress") or restaurant.get("address"),
        "Google Place ID": restaurant.get("gmaps_place_id"),
        "Restaurant Rating": details.get("rating") or restaurant.get("rating"),
        "User Rating Count": details.get("userRatingCount") or restaurant.get("user_rating_count"),
        "Top 5 Reviews Pulled": min(len(reviews), 5),
        "Top Dishes": top_dishes,
        "Distinct Dishes Found": len(dish_counter),
        "Google Maps URL": details.get("googleMapsUri", ""),
    }
    return summary_row, review_rows


def write_sheet(ws, rows: list[dict], title_fill: str) -> None:
    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)

    fill = PatternFill(fill_type="solid", fgColor=title_fill)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    for column_index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for column_cells in ws.iter_cols(
            min_col=column_index, max_col=column_index, min_row=2, max_row=ws.max_row
        ):
            for cell in column_cells:
                max_length = max(max_length, len(str(cell.value or "")))

        column_letter = get_column_letter(column_index)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 60)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"


def export_to_excel(path: Path, summary_rows: list[dict], review_rows: list[dict]) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Restaurant Summary"
    write_sheet(ws_summary, summary_rows, "1F4E78")

    ws_reviews = wb.create_sheet("Review Details")
    write_sheet(ws_reviews, review_rows, "2E7D32")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    args = parse_args()
    limit = args.limit if args.limit is not None else prompt_for_limit()
    output_path = resolve_output_path(args.output)

    restaurants = fetch_restaurants(limit)
    if not restaurants:
        print("No restaurants with Google Place IDs were found in the places table.")
        return

    summary_rows: list[dict] = []
    review_rows: list[dict] = []

    print(f"Processing {len(restaurants)} restaurant(s)...")
    for index, restaurant in enumerate(restaurants, start=1):
        place_id = restaurant["gmaps_place_id"]
        name = restaurant.get("name") or f"Restaurant {restaurant.get('id')}"
        print(f"[{index}/{len(restaurants)}] {name}")
        try:
            details = fetch_place_reviews(place_id)
            summary_row, restaurant_review_rows = build_restaurant_summary(restaurant, details)
        except requests.RequestException as exc:
            print(f"  Skipping due to Google Places API error: {exc}")
            summary_row = {
                "Restaurant ID": restaurant.get("id"),
                "Restaurant Name": name,
                "Address": restaurant.get("address"),
                "Google Place ID": place_id,
                "Restaurant Rating": restaurant.get("rating"),
                "User Rating Count": restaurant.get("user_rating_count"),
                "Top 5 Reviews Pulled": 0,
                "Top Dishes": "",
                "Distinct Dishes Found": 0,
                "Google Maps URL": "",
            }
            restaurant_review_rows = []

        summary_rows.append(summary_row)
        review_rows.extend(restaurant_review_rows)

    export_to_excel(output_path, summary_rows, review_rows)
    print(f"Saved Excel output to: {output_path}")


if __name__ == "__main__":
    main()
