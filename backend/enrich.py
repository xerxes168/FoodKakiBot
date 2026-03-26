"""
enrich.py
---------
Unified tagging and location-enrichment module for FoodKakiBot.

Changes in this version:
  - SUBCUISINE_TO_PARENT mapping: sub-cuisines (Ramen, Dim Sum, Mala, Hubei, etc.)
    now automatically include their parent cuisine tag (Japanese, Chinese, etc.)
  - Road-name tags: addresses are parsed to extract road names (e.g. "Orchard Road",
    "Arab Street") and stored as searchable tags.
  - expand_with_parent_cuisines(): applied after all tag-inference paths.
  - extract_road_name_from_address(): applied in LocationEnricher, GoogleEnricher,
    and AutoTagger so every enrichment path creates road-name tags.

PUBLIC API (importable by app.py and rag.py):
  auto_tags_from_google(details)          → list[str]
  expand_location_tags(tag, **opts)       → list[str]
  build_location_tag_sets(locs, b, c)     → list[list[str]]
  get_planning_area_from_postal(code)     → str | None
  MRT_STATIONS, PLANNING_AREA_NEIGHBORS,
  AREA_ALIAS_TO_PLANNING_AREA, POSTAL_DISTRICT_TO_AREA

CLI (run as script):
  python enrich.py location  --apply [--mrt-radius 600] [--use-onemap] [--skip-mrt] [--skip-area]
  python enrich.py google    --apply [--use-find] [--include-allergies] [--places-new] [--fields ...]
  python enrich.py auto      --apply [--source supabase|dataset] [--dataset PATH] [--sheet NAME]
                                     [--use-google-geocode]
  python enrich.py all       --apply   # runs location → auto → google in sequence

Common flags (all sub-commands):
  --apply            Write to Supabase (default: dry-run / preview)
  --limit N          Cap rows processed
  --report PATH      Write .csv / .json / .xlsx report
  --skip-if-tagged   Skip places that already have ≥1 tag  (google / auto)
"""

from __future__ import annotations

# ── stdlib ─────────────────────────────────────────────────────────────────────
import argparse
import ast
import csv
import json
import math
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

# ── optional deps ──────────────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*_, **__) -> bool:
        return False

try:
    from supabase import Client, create_client
except ImportError:
    Client = Any
    create_client = None

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = Workbook = None

# ══════════════════════════════════════════════════════════════════════════════
# §1  STATIC GEOGRAPHIC DATA
# ══════════════════════════════════════════════════════════════════════════════

MRT_STATIONS: list[tuple[str, float, float]] = [
    # North-South Line
    ("Jurong East", 1.3332, 103.7422), ("Bukit Batok", 1.3491, 103.7496),
    ("Bukit Gombak", 1.3587, 103.7518), ("Choa Chu Kang", 1.3854, 103.7443),
    ("Yew Tee", 1.3970, 103.7477), ("Kranji", 1.4252, 103.7619),
    ("Marsiling", 1.4327, 103.7742), ("Woodlands", 1.4370, 103.7864),
    ("Admiralty", 1.4408, 103.8006), ("Sembawang", 1.4490, 103.8199),
    ("Canberra", 1.4431, 103.8298), ("Yishun", 1.4293, 103.8353),
    ("Khatib", 1.4175, 103.8330), ("Yio Chu Kang", 1.3817, 103.8448),
    ("Ang Mo Kio", 1.3700, 103.8495), ("Bishan", 1.3510, 103.8485),
    ("Braddell", 1.3404, 103.8468), ("Toa Payoh", 1.3328, 103.8474),
    ("Novena", 1.3203, 103.8437), ("Newton", 1.3133, 103.8386),
    ("Orchard", 1.3044, 103.8319), ("Somerset", 1.3006, 103.8390),
    ("Dhoby Ghaut", 1.2990, 103.8455), ("City Hall", 1.2931, 103.8520),
    ("Raffles Place", 1.2835, 103.8514), ("Marina Bay", 1.2764, 103.8545),
    ("Marina South Pier", 1.2710, 103.8636),
    # East-West Line
    ("Pasir Ris", 1.3731, 103.9493), ("Tampines", 1.3540, 103.9454),
    ("Simei", 1.3432, 103.9530), ("Tanah Merah", 1.3273, 103.9462),
    ("Bedok", 1.3240, 103.9299), ("Kembangan", 1.3207, 103.9129),
    ("Eunos", 1.3196, 103.9032), ("Paya Lebar", 1.3180, 103.8923),
    ("Aljunied", 1.3162, 103.8832), ("Kallang", 1.3115, 103.8716),
    ("Lavender", 1.3073, 103.8634), ("Bugis", 1.3009, 103.8564),
    ("Tanjong Pagar", 1.2764, 103.8462), ("Outram Park", 1.2801, 103.8398),
    ("Tiong Bahru", 1.2863, 103.8269), ("Redhill", 1.2895, 103.8166),
    ("Queenstown", 1.2942, 103.8060), ("Commonwealth", 1.3021, 103.7983),
    ("Buona Vista", 1.3072, 103.7904), ("Dover", 1.3113, 103.7784),
    ("Clementi", 1.3152, 103.7649), ("Chinese Garden", 1.3424, 103.7323),
    ("Lakeside", 1.3444, 103.7209), ("Boon Lay", 1.3388, 103.7059),
    ("Pioneer", 1.3369, 103.6971), ("Joo Koon", 1.3278, 103.6783),
    ("Gul Circle", 1.3196, 103.6611), ("Tuas Crescent", 1.3212, 103.6490),
    ("Tuas West Road", 1.3294, 103.6395), ("Tuas Link", 1.3408, 103.6374),
    ("Expo", 1.3353, 103.9614), ("Changi Airport", 1.3572, 103.9884),
    # Circle Line
    ("Bras Basah", 1.2963, 103.8502), ("Esplanade", 1.2934, 103.8552),
    ("Promenade", 1.2934, 103.8610), ("Nicoll Highway", 1.2998, 103.8637),
    ("Stadium", 1.3064, 103.8751), ("Mountbatten", 1.3064, 103.8819),
    ("Dakota", 1.3088, 103.8882), ("MacPherson", 1.3264, 103.8897),
    ("Tai Seng", 1.3351, 103.8873), ("Bartley", 1.3424, 103.8798),
    ("Serangoon", 1.3499, 103.8732), ("Lorong Chuan", 1.3518, 103.8648),
    ("Marymount", 1.3490, 103.8393), ("Caldecott", 1.3373, 103.8392),
    ("Botanic Gardens", 1.3228, 103.8150), ("Farrer Road", 1.3175, 103.8075),
    ("Holland Village", 1.3119, 103.7964), ("one-north", 1.2990, 103.7878),
    ("Kent Ridge", 1.2929, 103.7841), ("Haw Par Villa", 1.2831, 103.7819),
    ("Pasir Panjang", 1.2764, 103.7919), ("Labrador Park", 1.2723, 103.8023),
    ("Telok Blangah", 1.2706, 103.8098), ("HarbourFront", 1.2655, 103.8218),
    ("Bayfront", 1.2828, 103.8591),
    # Downtown Line
    ("Bukit Panjang", 1.3784, 103.7763), ("Cashew", 1.3699, 103.7839),
    ("Hillview", 1.3629, 103.7671), ("Beauty World", 1.3412, 103.7756),
    ("King Albert Park", 1.3348, 103.7783), ("Sixth Avenue", 1.3302, 103.7964),
    ("Tan Kah Kee", 1.3261, 103.8075), ("Stevens", 1.3198, 103.8266),
    ("Little India", 1.3066, 103.8491), ("Rochor", 1.3044, 103.8523),
    ("Downtown", 1.2793, 103.8526), ("Telok Ayer", 1.2823, 103.8484),
    ("Chinatown", 1.2845, 103.8436), ("Fort Canning", 1.2923, 103.8438),
    ("Bencoolen", 1.2984, 103.8497), ("Jalan Besar", 1.3052, 103.8554),
    ("Bendemeer", 1.3139, 103.8622), ("Geylang Bahru", 1.3218, 103.8712),
    ("Mattar", 1.3271, 103.8828), ("Ubi", 1.3291, 103.8991),
    ("Kaki Bukit", 1.3352, 103.9091), ("Bedok North", 1.3344, 103.9192),
    ("Bedok Reservoir", 1.3368, 103.9321), ("Tampines West", 1.3468, 103.9384),
    ("Tampines East", 1.3575, 103.9531), ("Upper Changi", 1.3417, 103.9613),
    # Thomson-East Coast Line
    ("Woodlands North", 1.4479, 103.8200), ("Woodlands South", 1.4268, 103.7950),
    ("Springleaf", 1.3988, 103.8178), ("Lentor", 1.3887, 103.8349),
    ("Mayflower", 1.3757, 103.8388), ("Bright Hill", 1.3637, 103.8370),
    ("Upper Thomson", 1.3567, 103.8309), ("Napier", 1.3074, 103.8217),
    ("Orchard Boulevard", 1.3009, 103.8248), ("Great World", 1.2948, 103.8236),
    ("Havelock", 1.2891, 103.8354), ("Maxwell", 1.2800, 103.8444),
    ("Shenton Way", 1.2773, 103.8483), ("Gardens by the Bay", 1.2810, 103.8650),
    ("Tanjong Rhu", 1.2994, 103.8724), ("Katong Park", 1.3048, 103.8820),
    ("Tanjong Katong", 1.3032, 103.8980), ("Marine Parade", 1.3026, 103.9066),
    ("Marine Terrace", 1.3055, 103.9143), ("Siglap", 1.3097, 103.9272),
    ("Bayshore", 1.3151, 103.9400), ("Bedok South", 1.3212, 103.9450),
    ("Sungei Bedok", 1.3265, 103.9557),
    # North-East Line
    ("Clarke Quay", 1.2884, 103.8464), ("Farrer Park", 1.3123, 103.8542),
    ("Boon Keng", 1.3197, 103.8617), ("Potong Pasir", 1.3313, 103.8693),
    ("Woodleigh", 1.3392, 103.8710), ("Kovan", 1.3599, 103.8851),
    ("Hougang", 1.3712, 103.8920), ("Buangkok", 1.3832, 103.8928),
    ("Sengkang", 1.3916, 103.8954), ("Punggol", 1.4053, 103.9022),
    # JRL partial
    ("Gek Poh", 1.3445, 103.6958), ("Tawas", 1.3499, 103.6876),
]
_seen_mrt: set[str] = set()
MRT_STATIONS = [s for s in MRT_STATIONS if not (_seen_mrt.add(s[0]) or s[0] in _seen_mrt - {s[0]})]

PLANNING_AREA_NEIGHBORS: dict[str, list[str]] = {
    "Raffles Place":  ["Tanjong Pagar", "Chinatown", "City Hall", "Marina Bay", "Clarke Quay"],
    "Marina Bay":     ["Raffles Place", "City Hall", "Tanjong Pagar"],
    "Tanjong Pagar":  ["Raffles Place", "Chinatown", "Tiong Bahru", "Harbourfront"],
    "Chinatown":      ["Tanjong Pagar", "Raffles Place", "Clarke Quay", "Tiong Bahru"],
    "City Hall":      ["Raffles Place", "Marina Bay", "Bugis", "Clarke Quay"],
    "Clarke Quay":    ["Raffles Place", "City Hall", "Robertson Quay", "Chinatown"],
    "Robertson Quay": ["Clarke Quay", "Orchard", "Tiong Bahru", "Tanjong Pagar"],
    "Orchard":        ["Newton", "Tanglin", "Robertson Quay", "Novena"],
    "Tanglin":        ["Orchard", "Newton", "Holland Village", "Bukit Timah"],
    "Holland Village":["Tanglin", "Bukit Timah", "Queenstown", "Clementi"],
    "Bugis":          ["City Hall", "Kampong Glam", "Little India", "Kallang"],
    "Kampong Glam":   ["Bugis", "Little India", "Kallang"],
    "Little India":   ["Bugis", "Kampong Glam", "Novena", "Kallang"],
    "Newton":         ["Orchard", "Novena", "Tanglin", "Bukit Timah"],
    "Novena":         ["Newton", "Toa Payoh", "Balestier", "Bishan", "Bukit Timah"],
    "Balestier":      ["Novena", "Toa Payoh", "MacPherson", "Little India"],
    "Upper Thomson":  ["Novena", "Bishan", "Ang Mo Kio", "Mandai"],
    "Mandai":         ["Upper Thomson", "Woodlands", "Yishun", "Ang Mo Kio"],
    "Toa Payoh":      ["Novena", "Bishan", "Ang Mo Kio", "Kallang", "MacPherson"],
    "Bishan":         ["Ang Mo Kio", "Toa Payoh", "Novena", "Serangoon", "Upper Thomson"],
    "Kallang":        ["Bugis", "Geylang", "MacPherson", "Toa Payoh", "Marine Parade"],
    "MacPherson":     ["Kallang", "Geylang", "Toa Payoh", "Balestier", "Paya Lebar", "Potong Pasir"],
    "Potong Pasir":   ["MacPherson", "Serangoon", "Toa Payoh", "Geylang"],
    "Geylang":        ["Kallang", "MacPherson", "Paya Lebar", "Eunos", "Joo Chiat"],
    "Eunos":          ["Geylang", "Paya Lebar", "Kembangan", "Joo Chiat"],
    "Paya Lebar":     ["Geylang", "Serangoon", "Hougang", "Eunos"],
    "Joo Chiat":      ["Geylang", "Katong", "Eunos"],
    "Katong":         ["Joo Chiat", "Marine Parade", "Bedok"],
    "Marine Parade":  ["Katong", "Kallang", "Bedok"],
    "Bedok":          ["Tampines", "Pasir Ris", "Marine Parade", "Katong"],
    "Tampines":       ["Pasir Ris", "Bedok", "Changi", "Sengkang"],
    "Pasir Ris":      ["Tampines", "Changi", "Sengkang"],
    "Changi":         ["Tampines", "Pasir Ris", "Bedok"],
    "Serangoon":      ["Ang Mo Kio", "Bishan", "Hougang", "Sengkang", "Paya Lebar", "Potong Pasir"],
    "Hougang":        ["Ang Mo Kio", "Serangoon", "Sengkang", "Punggol"],
    "Sengkang":       ["Hougang", "Punggol", "Pasir Ris", "Serangoon"],
    "Punggol":        ["Hougang", "Sengkang"],
    "Ang Mo Kio":     ["Bishan", "Toa Payoh", "Serangoon", "Yio Chu Kang", "Hougang", "Upper Thomson"],
    "Yio Chu Kang":   ["Ang Mo Kio", "Hougang", "Seletar", "Serangoon"],
    "Seletar":        ["Yio Chu Kang", "Sengkang", "Ang Mo Kio"],
    "Woodlands":      ["Mandai", "Choa Chu Kang", "Bukit Panjang", "Lim Chu Kang", "Sembawang"],
    "Sembawang":      ["Woodlands", "Yishun"],
    "Yishun":         ["Mandai", "Sembawang", "Ang Mo Kio"],
    "Queenstown":     ["Tiong Bahru", "Clementi", "Buona Vista", "Holland Village", "Alexandra"],
    "Buona Vista":    ["Queenstown", "Clementi", "West Coast", "Pasir Panjang"],
    "Clementi":       ["Bukit Timah", "Bukit Batok", "Queenstown", "Buona Vista", "Jurong East"],
    "Tiong Bahru":    ["Chinatown", "Tanjong Pagar", "Queenstown", "Alexandra", "Robertson Quay"],
    "Alexandra":      ["Tiong Bahru", "Queenstown", "Harbourfront", "Buona Vista"],
    "Harbourfront":   ["Tanjong Pagar", "Alexandra", "Pasir Panjang"],
    "Pasir Panjang":  ["Harbourfront", "Buona Vista", "West Coast"],
    "West Coast":     ["Clementi", "Queenstown", "Buona Vista", "Pasir Panjang"],
    "Bukit Timah":    ["Newton", "Tanglin", "Holland Village", "Bukit Batok", "Bukit Panjang", "Clementi"],
    "Bukit Batok":    ["Bukit Timah", "Clementi", "Jurong East", "Choa Chu Kang", "Bukit Panjang"],
    "Bukit Panjang":  ["Choa Chu Kang", "Bukit Timah", "Bukit Batok", "Woodlands"],
    "Choa Chu Kang":  ["Bukit Batok", "Bukit Panjang", "Woodlands", "Tengah", "Jurong West"],
    "Jurong East":    ["Jurong West", "Bukit Batok", "Clementi"],
    "Jurong West":    ["Jurong East", "Choa Chu Kang", "Tuas"],
    "Tengah":         ["Bukit Batok", "Choa Chu Kang", "Jurong East"],
    "Tuas":           ["Jurong West"],
    "Lim Chu Kang":   ["Woodlands", "Choa Chu Kang"],
}

POSTAL_DISTRICT_TO_AREA: dict[str, str] = {
    "01": "Raffles Place",  "02": "Raffles Place",  "03": "Raffles Place",
    "04": "Marina Bay",     "05": "Marina Bay",     "06": "Chinatown",
    "07": "Tanjong Pagar",  "08": "Tanjong Pagar",
    "09": "Harbourfront",   "10": "Harbourfront",
    "11": "Buona Vista",    "12": "Pasir Panjang",  "13": "Pasir Panjang",
    "14": "Tiong Bahru",    "15": "Queenstown",     "16": "Alexandra",
    "17": "City Hall",      "18": "Bugis",          "19": "Bugis",
    "20": "Little India",   "21": "Little India",   "22": "Orchard",
    "23": "River Valley",   "24": "Tanglin",        "25": "Tanglin",
    "26": "Holland Village","27": "Holland Village","28": "Novena",
    "29": "Newton",         "30": "Newton",         "31": "Balestier",
    "32": "Toa Payoh",      "33": "Toa Payoh",      "34": "MacPherson",
    "35": "Kallang",        "36": "MacPherson",     "37": "Potong Pasir",
    "38": "Geylang",        "39": "Geylang",        "40": "Paya Lebar",
    "41": "Eunos",          "42": "Joo Chiat",      "43": "Katong",
    "44": "Marine Parade",  "45": "Marine Parade",  "46": "Bedok",
    "47": "Bedok",          "48": "Bedok",          "49": "Changi",
    "50": "Changi",         "51": "Tampines",       "52": "Pasir Ris",
    "53": "Hougang",        "54": "Serangoon",      "55": "Serangoon",
    "56": "Bishan",         "57": "Ang Mo Kio",     "58": "Clementi",
    "59": "Clementi",       "60": "Jurong East",    "61": "Jurong West",
    "62": "Jurong West",    "63": "Jurong West",    "64": "Jurong East",
    "65": "Bukit Batok",    "66": "Bukit Batok",    "67": "Choa Chu Kang",
    "68": "Bukit Panjang",  "69": "Tengah",         "70": "Lim Chu Kang",
    "71": "Lim Chu Kang",   "72": "Woodlands",      "73": "Woodlands",
    "75": "Yishun",         "76": "Sembawang",      "77": "Upper Thomson",
    "78": "Mandai",         "79": "Yio Chu Kang",   "80": "Seletar",
    "81": "Changi",         "82": "Sengkang",
}

AREA_ALIAS_TO_PLANNING_AREA: dict[str, str] = {
    "raffles place": "Raffles Place", "shenton way": "Raffles Place",
    "tanjong pagar plaza": "Tanjong Pagar", "robinson road": "Raffles Place",
    "battery road": "Raffles Place", "telok ayer": "Raffles Place",
    "amoy street": "Raffles Place", "club street": "Tanjong Pagar",
    "ann siang": "Tanjong Pagar", "tanjong pagar": "Tanjong Pagar",
    "chinatown": "Chinatown", "new bridge road": "Chinatown",
    "smith street": "Chinatown", "temple street": "Chinatown",
    "eu tong sen": "Chinatown", "marina bay sands": "Marina Bay",
    "marina bay": "Marina Bay", "marina centre": "Marina Bay",
    "suntec": "Marina Bay", "esplanade": "City Hall",
    "city hall": "City Hall", "boat quay": "Clarke Quay",
    "clarke quay": "Clarke Quay", "clark quay": "Clarke Quay",
    "robertson quay": "Robertson Quay", "river valley": "Robertson Quay",
    "fort canning": "Clarke Quay", "bras basah": "Bugis",
    "kampong glam": "Kampong Glam", "arab street": "Kampong Glam",
    "haji lane": "Kampong Glam", "beach road": "Bugis",
    "bugis": "Bugis", "golden mile": "Bugis",
    "little india": "Little India", "serangoon road": "Little India",
    "farrer park": "Little India", "mustafa": "Little India",
    "tekka": "Little India", "rochor": "Bugis",
    "orchard road": "Orchard", "orchard": "Orchard",
    "somerset": "Orchard", "cairnhill": "Orchard",
    "scotts road": "Orchard", "dhoby ghaut": "Orchard",
    "tanglin": "Tanglin", "dempsey": "Tanglin",
    "cuscaden": "Tanglin", "ardmore": "Tanglin",
    "nassim": "Tanglin", "stevens road": "Tanglin",
    "holland village": "Holland Village", "holland": "Holland Village",
    "holland road": "Holland Village", "ghim moh": "Holland Village",
    "farrer road": "Holland Village", "sixth avenue": "Holland Village",
    "king albert park": "Bukit Timah", "botanic gardens": "Holland Village",
    "bukit timah": "Bukit Timah", "beauty world": "Bukit Timah",
    "newton": "Newton", "novena": "Novena",
    "moulmein": "Novena", "thomson road": "Novena",
    "upper thomson": "Upper Thomson", "springleaf": "Upper Thomson",
    "lentor": "Upper Thomson", "thomson": "Upper Thomson",
    "mandai": "Mandai",
    "balestier": "Balestier", "boon keng": "Balestier",
    "bendemeer": "Balestier", "toa payoh": "Toa Payoh",
    "braddell": "Toa Payoh", "kallang": "Kallang",
    "lavender": "Kallang", "geylang bahru": "Kallang",
    "jalan besar": "Little India", "bishan": "Bishan",
    "marymount": "Bishan",
    "macpherson": "MacPherson", "mac pherson": "MacPherson",
    "potong pasir": "Potong Pasir", "ubi": "MacPherson",
    "geylang": "Geylang", "aljunied": "Geylang",
    "guillemard": "Geylang", "dakota": "Geylang",
    "eunos": "Eunos", "kembangan": "Eunos",
    "paya lebar": "Paya Lebar",
    "joo chiat": "Joo Chiat", "east coast road": "Joo Chiat",
    "tanjong katong": "Katong", "amber road": "Katong",
    "katong": "Katong", "marine parade": "Marine Parade",
    "marine drive": "Marine Parade", "east coast park": "Marine Parade",
    "east coast": "Marine Parade", "siglap": "Marine Parade",
    "mountbatten": "Marine Parade", "tanjong rhu": "Marine Parade",
    "bedok": "Bedok", "upper changi": "Bedok",
    "new upper changi": "Bedok", "chai chee": "Bedok",
    "tampines": "Tampines", "simei": "Tampines",
    "pasir ris": "Pasir Ris", "changi": "Changi",
    "loyang": "Changi", "changi airport": "Changi",
    "changi village": "Changi",
    "serangoon": "Serangoon", "kovan": "Serangoon",
    "lorong chuan": "Serangoon", "upper serangoon": "Serangoon",
    "hougang": "Hougang", "buangkok": "Hougang",
    "sengkang": "Sengkang", "punggol": "Punggol",
    "ang mo kio": "Ang Mo Kio", "amk": "Ang Mo Kio",
    "sin ming": "Bishan", "bright hill": "Bishan",
    "yio chu kang": "Yio Chu Kang", "seletar": "Seletar",
    "jalan kayu": "Seletar", "fernvale": "Sengkang",
    "queenstown": "Queenstown", "commonwealth": "Queenstown",
    "buona vista": "Buona Vista", "one-north": "Buona Vista",
    "one north": "Buona Vista", "kent ridge": "Buona Vista",
    "dover": "Buona Vista", "tiong bahru": "Tiong Bahru",
    "redhill": "Tiong Bahru", "havelock": "Tiong Bahru",
    "alexandra": "Alexandra", "bukit merah": "Alexandra",
    "harbourfront": "Harbourfront", "telok blangah": "Harbourfront",
    "sentosa": "Harbourfront", "vivocity": "Harbourfront",
    "pasir panjang": "Pasir Panjang", "haw par villa": "Pasir Panjang",
    "labrador": "Pasir Panjang", "west coast": "West Coast",
    "jurong east": "Jurong East", "jurong west": "Jurong West",
    "boon lay": "Jurong West", "pioneer": "Jurong West",
    "tuas": "Tuas", "clementi": "Clementi",
    "bukit batok": "Bukit Batok", "bukit gombak": "Bukit Batok",
    "choa chu kang": "Choa Chu Kang", "cck": "Choa Chu Kang",
    "yew tee": "Choa Chu Kang", "bukit panjang": "Bukit Panjang",
    "hillview": "Bukit Panjang", "cashew": "Bukit Panjang",
    "tengah": "Tengah", "lim chu kang": "Lim Chu Kang",
    "woodlands": "Woodlands", "marsiling": "Woodlands",
    "woodgrove": "Woodlands", "kranji": "Woodlands",
    "admiralty": "Sembawang", "sembawang": "Sembawang",
    "canberra": "Sembawang", "yishun": "Yishun", "khatib": "Yishun",
}


# ══════════════════════════════════════════════════════════════════════════════
# §2  TAG TAXONOMY DATA
# ══════════════════════════════════════════════════════════════════════════════

CUISINE_KEYWORDS: dict[str, Sequence[str]] = {
    "Japanese":   ("japanese","sushi","sashimi","ramen","udon","soba","yakitori",
                   "izakaya","tempura","tonkatsu","donburi","omakase","katsu","wagyu",
                   "gyoza","onigiri","teppanyaki","unagi","handroll","kaiseki"),
    "Korean":     ("korean","kimchi","bibimbap","bulgogi","tteokbokki","kbbq",
                   "korean bbq","banchan","samgyeopsal","jjajangmyeon","doenjang"),
    "Chinese":    ("chinese","dim sum","dumpling","wonton","char siew","cantonese",
                   "szechuan","sichuan","xiao long bao","claypot","zi char","tze char",
                   "hubei","hunan","yunnan","teochew","hokkien","shanghainese","dongbei"),
    "Mala":       ("mala","ma la","spicy pot","xiang guo","dry pot"),
    "Indian":     ("indian","biryani","briyani","tandoori","naan","masala","dosa",
                   "prata","roti prata","thosai","tikka","paneer","chaat"),
    "Malay":      ("malay","nasi lemak","satay","rendang","mee rebus","nasi padang",
                   "padang","laksa","otah","mee goreng"),
    "Thai":       ("thai","tom yum","pad thai","green curry","som tam","basil"),
    "Vietnamese": ("vietnamese","pho","banh mi","bun cha","spring roll"),
    "Italian":    ("italian","pasta","pizza","risotto","carbonara","lasagna",
                   "trattoria","gelato","tiramisu","focaccia"),
    "French":     ("french","confit","croissant","escargot","foie gras"),
    "Mexican":    ("mexican","taco","burrito","quesadilla","nachos","guacamole"),
    "Western":    ("western","steak","burger","grill","bistro","barbecue",
                   "fish and chips","brunch","roast"),
    "Seafood":    ("seafood","crab","lobster","oyster","prawn","clam","mussel","scallop"),
    "Vegetarian": ("vegetarian","vegan","plant-based","meat-free"),
    "Halal":      ("halal",),
    "Dessert":    ("dessert","cake","pastry","gelato","ice cream","sweet",
                   "brownie","tart","pudding","crepe","waffle","churros"),
    "Cafe":       ("cafe","coffee","latte","espresso","flat white","cappuccino","barista"),
    "Bubble Tea": ("bubble tea","boba","milk tea","gong cha","koi","liho","taro"),
    "Fast Food":  ("fast food","mcdonald","kfc","subway","burger king","jollibee"),
    "Singaporean":("hawker","kopitiam","chicken rice","char kway teow",
                   "bak kut teh","laksa","rojak","carrot cake","wanton mee"),
    "Taiwanese":  ("taiwanese","lu rou fan","braised pork rice","oyster omelette",
                   "beef noodle","scallion pancake","salted crispy chicken"),
    "Indonesian": ("indonesian","bakso","soto ayam","ayam penyet","bebek","mie ayam",
                   "gado gado","nasi campur bali"),
    "Mediterranean": ("mediterranean","mezze","falafel","grilled lamb","hummus platter"),
    "Middle Eastern": ("middle eastern","shawarma","kebab","hummus","labneh","manakish"),
    "Spanish":    ("spanish","tapas","paella","jamon","patatas bravas"),
    "Brunch":     ("brunch","eggs benedict","avocado toast","all day breakfast"),
}

ALLERGY_KEYWORDS: dict[str, Sequence[str]] = {
    "Gluten-Free":    ("gluten free","gluten-free","gf"),
    "Dairy-Free":     ("dairy free","dairy-free","lactose free","no dairy"),
    "Nut-Free":       ("nut free","nut-free","peanut-free","no nuts","tree nut"),
    "Shellfish-Free": ("shellfish free","shellfish-free","no shellfish"),
    "Egg-Free":       ("egg free","egg-free","no egg"),
    "Soy-Free":       ("soy free","soy-free","no soy"),
}

CUISINE_ALIASES: dict[str, str] = {
    "jpn": "Japanese", "jp": "Japanese", "korea": "Korean",
    "chinese food": "Chinese", "veg": "Vegetarian", "western food": "Western",
    "taiwan": "Taiwanese", "indo": "Indonesian",
    "mediterranean food": "Mediterranean",
}

# ── Specific sub-cuisine keyword detection ────────────────────────────────────
# Maps sub-cuisine tag names → their specific trigger keywords (lowercased).
# When ANY keyword is detected in place text, both the sub-cuisine tag AND
# its parent cuisine tag are assigned.
# e.g. text contains "ramen" → tags: ["Ramen", "Japanese"]
# e.g. text contains "hubei" → tags: ["Hubei", "Chinese"]
# e.g. text contains "dim sum" → tags: ["Dim Sum", "Chinese"]
SPECIFIC_CUISINE_KEYWORDS: dict[str, list[str]] = {
    # Japanese sub-types
    "Ramen":      ["ramen", "tsukemen", "tantanmen", "tonkotsu ramen"],
    "Sushi":      ["sushi", "sashimi", "omakase", "nigiri", "maki", "temaki", "kaiseki"],
    "Yakitori":   ["yakitori", "kushiyaki", "kushikatsu"],
    "Tempura":    ["tempura"],
    "Teppanyaki": ["teppanyaki"],
    # Chinese sub-types / regional cuisines
    "Dim Sum":    ["dim sum", "dimsum", "yum cha", "har gow", "siu mai", "char siu bao",
                   "cheong fun", "lo mai gai"],
    "Hotpot / Steamboat": ["hotpot", "hot pot", "steamboat", "shabu shabu",
                            "mookata", "chinese fondue"],
    "Mala":       ["mala", "ma la", "spicy pot", "xiang guo", "dry pot"],
    "Cantonese":  ["cantonese", "congee", "roast duck", "wonton noodles", "poon choi",
                   "char siu rice"],
    "Teochew":    ["teochew", "chaozhou", "braised duck teochew"],
    "Hokkien":    ["hokkien"],
    "Sichuan":    ["sichuan", "szechuan", "mapo tofu", "kung pao", "dan dan noodles"],
    "Hubei":      ["hubei"],
    "Hunan":      ["hunan", "hunanese"],
    "Yunnan":     ["yunnan", "crossing the bridge noodles"],
    "Shanghainese": ["shanghainese", "xiao long bao", "soup dumpling", "xlb",
                     "red braised pork"],
    "Xinjiang":   ["xinjiang", "uyghur", "lamb skewer xinjiang"],
    "Dongbei":    ["dongbei", "northeast chinese"],
    # Korean sub-types
    "Korean BBQ": ["kbbq", "korean bbq", "samgyeopsal", "galbi", "korean grill",
                   "korean barbecue"],
    # Indian sub-types
    "Biryani":    ["biryani", "briyani", "dum biryani"],
    "Prata":      ["prata", "roti prata", "canai"],
    "Dosa":       ["dosa", "thosai", "uttapam", "idli"],
    # Italian sub-types
    "Pizza":      ["pizza", "pizzeria", "neapolitan pizza", "roman pizza"],
    # Western sub-types
    "Burgers":    ["burger", "smash burger", "wagyu burger", "cheeseburger"],
    "Steakhouse": ["steakhouse", "steak house", "chophouse", "prime rib",
                   "dry-aged steak"],
    "BBQ":        ["smokehouse", "brisket", "pulled pork", "american bbq",
                   "bbq ribs"],
    "Sandwiches": ["sandwich shop", "sub shop", "deli sandwich"],
    "Taiwanese":  ["taiwanese", "lu rou fan", "braised pork rice", "oyster omelette",
                   "beef noodle soup", "scallion pancake", "salted crispy chicken"],
    "Indonesian": ["indonesian", "nasi padang", "ayam penyet", "bakso", "gado gado",
                   "mie ayam", "soto ayam"],
    "Middle Eastern": ["middle eastern", "shawarma", "kebab", "hummus", "labneh",
                       "manakish", "falafel wrap"],
    "Mediterranean": ["mediterranean", "mezze", "falafel", "gyro", "grilled lamb"],
    "Spanish":    ["spanish", "tapas", "paella", "jamon", "patatas bravas"],
}

# Maps each sub-cuisine tag to its parent cuisine tag.
SUBCUISINE_PARENT: dict[str, str] = {
    "Ramen": "Japanese", "Sushi": "Japanese", "Yakitori": "Japanese",
    "Tempura": "Japanese", "Teppanyaki": "Japanese",
    "Dim Sum": "Chinese", "Hotpot / Steamboat": "Chinese", "Mala": "Chinese",
    "Cantonese": "Chinese", "Teochew": "Chinese", "Hokkien": "Chinese",
    "Sichuan": "Chinese", "Hubei": "Chinese", "Hunan": "Chinese",
    "Yunnan": "Chinese", "Shanghainese": "Chinese", "Xinjiang": "Chinese",
    "Dongbei": "Chinese",
    "Korean BBQ": "Korean",
    "Biryani": "Indian", "Prata": "Indian", "Dosa": "Indian",
    "Pizza": "Italian",
    "Burgers": "Western", "Steakhouse": "Western", "BBQ": "Western",
    "Sandwiches": "Western",
    "Taiwanese": "Taiwanese", "Indonesian": "Indonesian",
    "Mediterranean": "Mediterranean", "Middle Eastern": "Middle Eastern",
    "Spanish": "Spanish",
}

# Google primaryType → cuisine label
PRIMARY_TYPE_TO_CUISINE: dict[str, str] = {
    "japanese_restaurant": "Japanese", "korean_restaurant": "Korean",
    "chinese_restaurant": "Chinese", "indian_restaurant": "Indian",
    "thai_restaurant": "Thai", "vietnamese_restaurant": "Vietnamese",
    "italian_restaurant": "Italian", "french_restaurant": "French",
    "mexican_restaurant": "Mexican", "seafood_restaurant": "Seafood",
    "vegetarian_restaurant": "Vegetarian", "vegan_restaurant": "Vegetarian",
    "halal_restaurant": "Halal", "cafe": "Cafe", "coffee_shop": "Cafe",
    "dessert_shop": "Dessert", "ice_cream_shop": "Dessert",
    "bakery": "Bakery", "steak_house": "Steakhouse", "american_restaurant": "Western",
    "barbecue_restaurant": "BBQ", "malay_restaurant": "Malay",
    "singaporean_restaurant": "Singaporean", "hot_pot_restaurant": "Hotpot / Steamboat",
    "ramen_restaurant": "Ramen", "sushi_restaurant": "Sushi",
    "dim_sum_restaurant": "Dim Sum", "bubble_tea_store": "Bubble Tea",
    "fast_food_restaurant": "Fast Food", "pizza_restaurant": "Pizza",
    "sandwich_shop": "Sandwiches", "buffet_restaurant": "Buffet",
    "taiwanese_restaurant": "Taiwanese", "indonesian_restaurant": "Indonesian",
    "mediterranean_restaurant": "Mediterranean", "middle_eastern_restaurant": "Middle Eastern",
    "spanish_restaurant": "Spanish", "brunch_restaurant": "Brunch",
    "ramen_restaurant": "Ramen", "sushi_restaurant": "Sushi",
    "steak_house": "Steakhouse", "barbecue_restaurant": "BBQ",
    "kebab_shop": "Middle Eastern", "falafel_restaurant": "Middle Eastern",
    "breakfast_restaurant": "Brunch",
}

# Google type → non-cuisine attribute tag
GOOGLE_TYPE_TO_ATTR_TAG: dict[str, str] = {
    "meal_takeaway": "Takeaway", "meal_delivery": "Delivery",
    "fast_food_restaurant": "Fast Food", "buffet_restaurant": "Buffet",
    "brunch_restaurant": "Brunch", "cafe": "Cafe", "bar": "Bar",
    "wine_bar": "Wine Bar", "bakery": "Bakery", "dessert_shop": "Dessert",
    "ice_cream_shop": "Ice Cream", "bubble_tea_store": "Bubble Tea",
    "juice_shop": "Juice Bar", "tea_house": "Tea House", "diner": "Diner",
    "food_court": "Food Court", "night_club": "Nightclub",
    "vegan_restaurant": "Vegetarian", "vegetarian_restaurant": "Vegetarian",
}

NON_SAVORY_VENUE_TYPES: set[str] = {
    "bakery", "cafe", "coffee_shop", "dessert_shop", "ice_cream_shop",
    "bubble_tea_store", "juice_shop", "tea_house",
}

NON_SAVORY_CUISINE_TAGS: set[str] = {
    "Bakery", "Cafe", "Dessert", "Ice Cream", "Bubble Tea", "Juice Bar",
    "Tea House", "Brunch",
}

_LEGACY_TYPE_TO_TAG: dict[str, list[str]] = {
    "bakery": ["Dessert"], "cafe": ["Cafe", "Dessert"],
    "meal_takeaway": ["Fast Food"], "meal_delivery": ["Fast Food"],
    "bar": ["Bar"], "night_club": ["Bar"],
}

NON_ASCII_RE = re.compile(r"[^\x00-\x7F]")
PRICE_NUMBER_RE = re.compile(r"\$(\d+(?:\.\d{1,2})?)")
_POSTAL_RE = re.compile(r"\b(S|s)?(\d{6})\b")

PRICE_LEVEL_TO_TAG = {0: "Free", 1: "Budget", 2: "Mid-Range", 3: "Expensive", 4: "Premium"}
AREA_COMPONENT_TYPES = (
    "neighborhood", "sublocality_level_1", "sublocality",
    "locality", "administrative_area_level_2",
)

# ── Road name extraction ───────────────────────────────────────────────────────
# Matches Singapore road names like "Orchard Road", "Smith Street 45",
# "Arab Street", "Tanjong Pagar Road", "Harbourfront Walk".
_ROAD_SUFFIX_PAT = (
    r"Road|Street|Avenue|Ave|Drive|Lane|Walk|Way|Place|Close|Crescent"
    r"|Boulevard|Link|Terrace|Rise|View|Grove|Hill|Park|Gardens|Quay"
    r"|Square|Loop|Court|Alley|Promenade|Esplanade|Rd|St|Dr|Ln|Blvd"
)
_ROAD_NAME_RE = re.compile(
    r"(?:"
    r"(?:^|,)\s*"                          # start or after comma
    r"(?:Blk\s+\d+[A-Za-z]?\s*,?\s*)?"   # optional "Blk 123," OR "Blk 123 " (no comma)
    r"(?:#[\w/\\-]+\s*,?\s*)?"            # optional "#01-23,"
    r"(?:\d+[A-Za-z]?\s+)?"              # optional house number "68 "
    r")"
    r"([A-Z][A-Za-z \'-]{2,50}"
    r"(?:" + _ROAD_SUFFIX_PAT + r")"
    r"(?:\s+\d+)?)",                       # optional trailing number e.g. "Street 45"
    re.IGNORECASE,
)


# ══════════════════════════════════════════════════════════════════════════════
# §3  SHARED UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def normalize_address_text(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_tag_name(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def is_english_tag(value: str) -> bool:
    c = normalize_tag_name(value)
    return bool(c and not NON_ASCII_RE.search(c) and re.search(r"[A-Za-z]", c))


def sanitize_area_candidate(value: str) -> Optional[str]:
    c = normalize_tag_name(value)
    if not c or not is_english_tag(c):
        return None
    if c.lower() in {"singapore", "sg"}:
        return None
    if re.search(r"\d", c) or "," in c or len(c) > 40:
        return None
    return c


def to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def extract_postal_code(text: str) -> Optional[str]:
    if not text:
        return None
    m = _POSTAL_RE.search(text)
    return m.group(2) if m else None


def haversine_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    R = 6_371_000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    a = math.sin(math.radians(lat2 - lat1) / 2) ** 2 + \
        math.cos(p1) * math.cos(p2) * math.sin(math.radians(lng2 - lng1) / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def chunked(items: Sequence[Any], size: int) -> Iterable[Sequence[Any]]:
    for i in range(0, len(items), size):
        yield items[i: i + size]


def safe_json_loads(raw: Any) -> Any:
    if raw is None:
        return None
    if isinstance(raw, (dict, list)):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except Exception:
        pass
    try:
        return ast.literal_eval(text)
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# §4  LOCATION HELPERS  (public API — imported by app.py)
# ══════════════════════════════════════════════════════════════════════════════

def get_planning_area_from_postal(postal_code: str) -> Optional[str]:
    code = (postal_code or "").strip().lstrip("Ss")
    return POSTAL_DISTRICT_TO_AREA.get(code[:2]) if len(code) >= 2 else None


def _canonical_planning_area(tag: str) -> Optional[str]:
    norm = normalize_text(tag)
    if norm.startswith("near ") and "mrt" in norm:
        return None
    for pa in PLANNING_AREA_NEIGHBORS:
        if normalize_text(pa) == norm:
            return pa
    return AREA_ALIAS_TO_PLANNING_AREA.get(norm)


def expand_location_tags(
    location_tag: str,
    *,
    include_neighbors: bool = True,
    max_neighbors: int = 4,
) -> list[str]:
    if not location_tag:
        return []
    norm = normalize_text(location_tag)
    if norm.startswith("near ") and "mrt" in norm:
        return [location_tag]
    canonical = _canonical_planning_area(location_tag)
    if not canonical:
        return [location_tag]
    result = [canonical]
    if include_neighbors:
        result.extend(PLANNING_AREA_NEIGHBORS.get(canonical, [])[:max_neighbors])
    seen: set[str] = set()
    return [t for t in result if not (t in seen or seen.add(t))]


def build_location_tag_sets(
    location_tags: list[str],
    budget_tags: list[str],
    cuisine_tags: list[str],
    *,
    max_neighbors: int = 3,
) -> list[list[str]]:
    if not location_tags:
        return []
    expanded: list[str] = []
    for lt in location_tags:
        for tag in expand_location_tags(lt, max_neighbors=max_neighbors):
            if tag not in expanded:
                expanded.append(tag)
    qualifiers = budget_tags + cuisine_tags
    if not qualifiers:
        return [[loc] for loc in expanded]
    return [[loc, q] for loc in expanded for q in qualifiers]


def area_from_address(address: str) -> Optional[str]:
    if not address:
        return None
    road = extract_road_name_from_address(address)
    if road:
        direct = AREA_ALIAS_TO_PLANNING_AREA.get(normalize_text(road))
        if direct:
            return direct

    norm = normalize_address_text(address)
    for phrase in sorted(AREA_ALIAS_TO_PLANNING_AREA, key=len, reverse=True):
        norm_phrase = normalize_address_text(phrase)
        if not norm_phrase:
            continue
        if re.search(rf"(?<![a-z0-9]){re.escape(norm_phrase)}(?![a-z0-9])", norm):
            return AREA_ALIAS_TO_PLANNING_AREA[phrase]

    postal = extract_postal_code(address)
    if postal:
        return get_planning_area_from_postal(postal)
    return None


def nearby_mrt_tags(lat: float, lng: float, radius_m: float) -> list[str]:
    return [f"Near {name} MRT" for name, slat, slng in MRT_STATIONS
            if haversine_m(lat, lng, slat, slng) <= radius_m]


def extract_road_name_from_address(address: str) -> Optional[str]:
    """
    Extract a capitalised road/street name from a Singapore address string
    and return it as a searchable tag — e.g. "Orchard Road", "Arab Street",
    "Tanjong Pagar Road", "Harbourfront Walk".

    Returns None when no road name can be confidently identified.
    """
    if not address:
        return None
    # Strip noise: unit numbers, postal codes, country name
    clean = re.sub(r"#[\w/\\-]+\s*,?\s*", " ", address)
    clean = re.sub(r",?\s*(?:Singapore\s*)?\b\d{6}\b", "", clean, flags=re.IGNORECASE)
    clean = re.sub(r"\bS\d{6}\b", "", clean)
    clean = re.sub(r"\bSingapore\b", "", clean, flags=re.IGNORECASE)
    clean = re.sub(r"\s{2,}", " ", clean).strip()

    m = _ROAD_NAME_RE.search(clean)
    if not m:
        return None

    road = m.group(1).strip().strip(",").strip()
    # Strip any leading house/block number that slipped through
    road = re.sub(r"^\d+[A-Za-z]?\s+", "", road)
    # Title-case each word for consistency with other tags
    road = " ".join(w.capitalize() for w in road.split())
    return road if len(road) >= 5 and re.search(r"[A-Za-z]", road) else None


# ══════════════════════════════════════════════════════════════════════════════
# §5  TAG INFERENCE  (public API — auto_tags_from_google imported by app.py)
# ══════════════════════════════════════════════════════════════════════════════

def extract_reviews_text(raw_reviews: Any) -> str:
    parsed = safe_json_loads(raw_reviews)
    if not parsed:
        return ""
    if isinstance(parsed, dict) and "reviews" in parsed:
        parsed = parsed["reviews"]
    texts = []
    if isinstance(parsed, list):
        for item in parsed:
            t = (item.get("text", "") if isinstance(item, dict) else item) or ""
            if str(t).strip():
                texts.append(str(t).strip())
    return "\n".join(texts)


def extract_google_types(details: dict) -> list[str]:
    values: list[str] = []
    primary = details.get("primaryType")
    if isinstance(primary, str) and primary.strip():
        values.append(primary.strip())
    for item in details.get("types") or []:
        if isinstance(item, str) and item.strip():
            values.append(item.strip())
    seen: set[str] = set()
    deduped: list[str] = []
    for value in values:
        key = normalize_text(value).replace(" ", "_")
        if key in seen:
            continue
        seen.add(key)
        deduped.append(value)
    return deduped


def extract_google_display_name(details: dict) -> str:
    display = details.get("displayName")
    if isinstance(display, dict):
        text = str(display.get("text") or "").strip()
        if text:
            return text
    return str(details.get("name") or "").strip()


def extract_address_text_from_details(details: dict) -> str:
    parts: list[str] = []
    for key in ("formattedAddress", "shortFormattedAddress", "formatted_address"):
        value = str(details.get(key) or "").strip()
        if value:
            parts.append(value)
    postal = details.get("postalAddress")
    if isinstance(postal, dict):
        lines = postal.get("addressLines") or []
        if isinstance(lines, list):
            for line in lines:
                text = str(line or "").strip()
                if text:
                    parts.append(text)
        for key in ("locality", "administrativeArea", "postalCode"):
            value = str(postal.get(key) or "").strip()
            if value:
                parts.append(value)
    return ", ".join(parts)


def infer_specific_cuisine_tags(text: str) -> list[str]:
    """
    Scan text for specific sub-cuisine keywords and return BOTH the sub-cuisine
    tag AND its parent cuisine tag.

    Examples:
      "tonkotsu ramen" → ["Ramen", "Japanese"]
      "dim sum yum cha" → ["Dim Sum", "Chinese"]
      "hubei cuisine"   → ["Hubei", "Chinese"]
      "sichuan mala hotpot" → ["Sichuan", "Chinese", "Mala", "Hotpot / Steamboat"]

    This is the correct replacement for the earlier SUBCUISINE_TO_PARENT approach,
    which never fired because sub-cuisine tag names were never produced by the
    keyword-scoring path — only parent cuisine names were.
    """
    norm = re.sub(r"\s+", " ", (text or "").strip().lower())
    found_sub: list[str] = []
    seen_sub: set[str] = set()

    for sub_cuisine, keywords in SPECIFIC_CUISINE_KEYWORDS.items():
        for kw in keywords:
            kw_norm = kw.lower().strip()
            # Use word-boundary matching for short keywords to avoid false positives
            if len(kw_norm) <= 5:
                matched = bool(re.search(r"\b" + re.escape(kw_norm) + r"\b", norm))
            else:
                matched = kw_norm in norm
            if matched:
                if sub_cuisine not in seen_sub:
                    seen_sub.add(sub_cuisine)
                    found_sub.append(sub_cuisine)
                break  # only need one keyword match per sub-cuisine

    # Add parent cuisine tags for all detected sub-cuisines
    result = list(found_sub)
    parents_seen: set[str] = {s.lower() for s in found_sub}
    for sub in found_sub:
        parent = SUBCUISINE_PARENT.get(sub)
        if parent and parent.lower() not in parents_seen:
            parents_seen.add(parent.lower())
            result.append(parent)

    return result


def expand_with_parent_cuisines(tags: list[str]) -> list[str]:
    """
    Legacy helper kept for the Google-API path where PRIMARY_TYPE_TO_CUISINE
    can produce sub-cuisine tag names like "Ramen", "Dim Sum", "Sushi" etc.
    In that case this function adds the parent tag as well.

    For text-inferred tags the correct entry point is infer_specific_cuisine_tags().
    """
    seen: set[str] = {t.lower() for t in tags}
    result = list(tags)
    for tag in list(tags):
        parent = SUBCUISINE_PARENT.get(tag)
        if parent and parent.lower() not in seen:
            seen.add(parent.lower())
            result.append(parent)
    return result


def infer_cuisine_tags(
    text: str,
    label_name: str = "",
    *,
    min_score: int = 2,
    max_tags: int = 3,
) -> list[str]:
    """
    Infer cuisine tags from free text.

    Two-pass approach:
      1. Broad scoring against CUISINE_KEYWORDS → top-level parent cuisines
         (Japanese, Chinese, Korean, …)
      2. Specific keyword scan via infer_specific_cuisine_tags → sub-cuisine tags
         AND their parent (Ramen+Japanese, Dim Sum+Chinese, Hubei+Chinese, …)

    The two result sets are merged and deduplicated so we never double-assign
    a parent tag that was already found by either pass.
    """
    score: dict[str, int] = {}
    norm_text = normalize_text(text)
    norm_name = normalize_text(label_name)
    canonical_label = CUISINE_ALIASES.get(norm_name, label_name.strip())
    if canonical_label in CUISINE_KEYWORDS:
        score[canonical_label] = score.get(canonical_label, 0) + 4
    for cuisine, keywords in CUISINE_KEYWORDS.items():
        for kw in keywords:
            kw_norm = normalize_text(kw)
            if not kw_norm:
                continue
            matched_in_text = kw_norm in norm_text
            matched_in_name = bool(norm_name and kw_norm in norm_name)
            if matched_in_text:
                score[cuisine] = score.get(cuisine, 0) + (2 if " " in kw_norm else 1)
            # Place-name matches are a stronger signal than generic text.
            if matched_in_name:
                score[cuisine] = score.get(cuisine, 0) + (3 if " " in kw_norm else 2)
    ranked = sorted(score.items(), key=lambda p: (-p[1], p[0]))
    base = [c for c, pts in ranked if pts >= min_score and is_english_tag(c)][:max_tags]

    # Second pass: detect sub-cuisine keywords → adds e.g. Ramen+Japanese,
    # Dim Sum+Chinese, Hubei+Chinese, Sichuan+Chinese, Mala+Chinese, etc.
    specific = infer_specific_cuisine_tags(text)

    # Merge: start from specific (more detailed), then append base tags not already present
    seen: set[str] = {t.lower() for t in specific}
    merged = list(specific)
    for t in base:
        if t.lower() not in seen:
            seen.add(t.lower())
            merged.append(t)
    return merged


def infer_cuisine_tags_from_types(types: Sequence[str]) -> list[str]:
    seen: set[str] = set()
    out = []
    for t in types:
        label = PRIMARY_TYPE_TO_CUISINE.get(normalize_text(t).replace(" ", "_"))
        if label and label not in seen:
            seen.add(label)
            out.append(label)
    # Automatically include parent cuisine tags (e.g. Ramen → Japanese)
    return expand_with_parent_cuisines(out)


def infer_attribute_tags_from_types(types: Sequence[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for t in types:
        label = GOOGLE_TYPE_TO_ATTR_TAG.get(normalize_text(t).replace(" ", "_"))
        if label and label not in seen:
            seen.add(label)
            out.append(label)
    return out


def name_has_non_savory_signal(name: str) -> bool:
    norm_name = normalize_text(name)
    if not norm_name:
        return False
    for cuisine, keywords in CUISINE_KEYWORDS.items():
        if cuisine not in NON_SAVORY_CUISINE_TAGS:
            continue
        for kw in keywords:
            kw_norm = normalize_text(kw)
            if kw_norm and kw_norm in norm_name:
                return True
    return False


def filter_cuisine_tags_for_venue(
    types: Sequence[str],
    cuisine_tags: Sequence[str],
    *,
    place_name: str = "",
) -> list[str]:
    type_keys = {normalize_text(t).replace(" ", "_") for t in types}
    has_non_savory_signal = bool(type_keys.intersection(NON_SAVORY_VENUE_TYPES)) or name_has_non_savory_signal(place_name)
    if not has_non_savory_signal:
        return list(cuisine_tags)

    explicit_savory_type = any(
        key.endswith("_restaurant") and key not in NON_SAVORY_VENUE_TYPES
        for key in type_keys
    )
    if explicit_savory_type:
        return list(cuisine_tags)

    allowed_from_name = set(infer_cuisine_tags(place_name, place_name))
    filtered = [
        tag for tag in cuisine_tags
        if tag in NON_SAVORY_CUISINE_TAGS or tag in allowed_from_name
    ]

    if filtered:
        seen: set[str] = set()
        return [tag for tag in filtered if not (tag.lower() in seen or seen.add(tag.lower()))]
    return list(allowed_from_name)


def infer_allergy_tags(text: str) -> list[str]:
    norm = normalize_text(text)
    return [tag for tag, kws in ALLERGY_KEYWORDS.items()
            if any(normalize_text(kw) in norm for kw in kws)]


def normalize_price_level(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, str):
        mapping = {
            "PRICE_LEVEL_FREE": 0, "PRICE_LEVEL_INEXPENSIVE": 1,
            "PRICE_LEVEL_MODERATE": 2, "PRICE_LEVEL_EXPENSIVE": 3,
            "PRICE_LEVEL_VERY_EXPENSIVE": 4,
        }
        key = value.strip().upper()
        if key in mapping:
            return mapping[key]
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def infer_price_range_tag(price_level: Any, text: str = "") -> str:
    level = normalize_price_level(price_level)
    if level is not None:
        return PRICE_LEVEL_TO_TAG.get(level, "Mid-Range")
    norm = normalize_text(text)
    if any(t in norm for t in ("cheap", "affordable", "budget", "value for money")):
        return "Budget"
    if any(t in norm for t in ("expensive", "pricey", "premium", "fine dining", "$$$")):
        return "Expensive"
    if "$$" in norm:
        return "Mid-Range"
    amounts = [float(m.group(1)) for m in PRICE_NUMBER_RE.finditer(text)]
    if amounts:
        avg = sum(amounts) / len(amounts)
        return "Budget" if avg <= 15 else ("Mid-Range" if avg <= 35 else "Expensive")
    return "Mid-Range"


def _extract_area_from_address_components(
    components: list[dict],
    long_key: str = "long_name",
) -> Optional[str]:
    for wanted in AREA_COMPONENT_TYPES:
        for comp in components:
            if wanted in set(comp.get("types", [])):
                name = str(comp.get(long_key) or comp.get("longText") or "").strip()
                if name:
                    return name
    return None


def infer_area_tag_from_details(details: dict) -> Optional[str]:
    comps = details.get("address_components") or details.get("addressComponents")
    if isinstance(comps, list):
        key = "long_name" if details.get("address_components") else "longText"
        raw = _extract_area_from_address_components(comps, key)
        clean = sanitize_area_candidate(raw or "")
        if clean:
            return clean

    address_area = area_from_address(extract_address_text_from_details(details))
    if address_area:
        return sanitize_area_candidate(address_area)

    postal = extract_postal_code(extract_address_text_from_details(details) or "")
    if postal:
        return sanitize_area_candidate(get_planning_area_from_postal(postal) or "")

    return None


def infer_location_tags_from_name(name: str) -> list[str]:
    """
    Infer Singapore area/location tags from a restaurant name by scanning for
    known area aliases and planning area names.

    Examples:
      "Orchard Ramen"             → ["Orchard"]
      "Bugis Street Noodles"      → ["Bugis"]
      "Tanjong Pagar Chicken Rice"→ ["Tanjong Pagar"]
      "Tiong Bahru Bakery"        → ["Tiong Bahru"]
    """
    if not name:
        return []
    norm = normalize_text(name)
    found: list[str] = []
    seen: set[str] = set()

    # Sort aliases longest-first so "tanjong pagar" matches before "pagar"
    for alias, area in sorted(AREA_ALIAS_TO_PLANNING_AREA.items(), key=lambda x: -len(x[0])):
        if not alias:
            continue
        if re.search(r"(?<![a-z])" + re.escape(alias) + r"(?![a-z])", norm):
            if area.lower() not in seen:
                seen.add(area.lower())
                found.append(area)

    # Also try road-name extraction from the name itself (e.g. "Keong Saik Road Kitchen")
    road = extract_road_name_from_address(name)
    if road:
        area_from_road = AREA_ALIAS_TO_PLANNING_AREA.get(normalize_text(road))
        if area_from_road and area_from_road.lower() not in seen:
            seen.add(area_from_road.lower())
            found.append(area_from_road)
        elif not area_from_road and road.lower() not in seen:
            seen.add(road.lower())
            found.append(road)

    return found


def auto_tags_from_google(details: dict) -> list[str]:
    """
    Public API — imported by app.py.
    Fast heuristic tag extraction from a Google Places Details result.
    Includes parent cuisine expansion automatically.
    """
    tags: set[str] = set()
    google_types = extract_google_types(details)
    for t in google_types:
        key = normalize_text(t).replace(" ", "_")
        tags.update(_LEGACY_TYPE_TO_TAG.get(key, []))
        attr = GOOGLE_TYPE_TO_ATTR_TAG.get(key)
        if attr:
            tags.add(attr)
        cuisine = PRIMARY_TYPE_TO_CUISINE.get(key)
        if cuisine:
            tags.add(cuisine)
    place_name = extract_google_display_name(details)
    text = " ".join(filter(None, [
        place_name,
        " ".join(google_types),
        extract_address_text_from_details(details),
    ]))
    text_cuisines = infer_cuisine_tags(text, place_name)
    type_cuisines = [tag for tag in tags if tag in CUISINE_KEYWORDS or tag in SUBCUISINE_PARENT]
    merged = type_cuisines + [tag for tag in text_cuisines if tag not in type_cuisines]
    tags.update(filter_cuisine_tags_for_venue(google_types, merged, place_name=place_name))
    # Location tags inferred from the restaurant name itself
    tags.update(infer_location_tags_from_name(place_name))
    # Expand sub-cuisines to include parent cuisines
    return sorted(expand_with_parent_cuisines(list(tags)))


# ══════════════════════════════════════════════════════════════════════════════
# §6  SUPABASE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def fetch_all_rows(
    supabase: Client,
    table: str,
    select: str = "*",
    page_size: int = 1000,
) -> list[dict]:
    rows: list[dict] = []
    offset = 0
    while True:
        batch = (supabase.table(table).select(select)
                 .range(offset, offset + page_size - 1).execute().data or [])
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return rows


def fetch_tags_map(supabase: Client) -> dict[str, int]:
    rows = fetch_all_rows(supabase, "tags", "id, name")
    return {row["name"].lower(): row["id"] for row in rows if row.get("name")}


def fetch_existing_links(supabase: Client, place_ids: list[int]) -> set[tuple[int, int]]:
    pairs: set[tuple[int, int]] = set()
    for batch in chunked(place_ids, 500):
        rows = (supabase.table("place_tags").select("place_id, tag_id")
                .in_("place_id", list(batch)).execute().data or [])
        for row in rows:
            pairs.add((row["place_id"], row["tag_id"]))
    return pairs


def ensure_tag(
    supabase: Client,
    name: str,
    tags_map: dict[str, int],
    *,
    apply: bool,
    _synthetic_counter: list[int],
) -> Optional[int]:
    clean = normalize_tag_name(name)
    if not clean:
        return None
    key = clean.lower()
    if key in tags_map:
        return tags_map[key]
    if not apply:
        fake = _synthetic_counter[0]
        _synthetic_counter[0] -= 1
        tags_map[key] = fake
        return fake
    try:
        resp = supabase.table("tags").insert({"name": clean}).execute()
        new_id = (resp.data or [{}])[0].get("id")
        if new_id:
            tags_map[key] = new_id
            return new_id
    except Exception:
        pass
    res2 = supabase.table("tags").select("id").ilike("name", clean).limit(1).execute()
    row = (res2.data or [None])[0]
    if row and row.get("id"):
        tags_map[key] = row["id"]
        return row["id"]
    return None


def insert_links_batch(supabase: Client, links: list[dict], *, apply: bool) -> int:
    if not apply or not links:
        return 0
    inserted = 0
    for batch in chunked(links, 250):
        try:
            supabase.table("place_tags").insert(list(batch)).execute()
            inserted += len(batch)
        except Exception:
            for row in batch:
                try:
                    supabase.table("place_tags").insert(row).execute()
                    inserted += 1
                except Exception:
                    pass
    return inserted


def write_report(path: str, rows: list[dict], fieldnames: list[str]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if p.suffix.lower() == ".csv":
        with p.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
    elif p.suffix.lower() == ".xlsx":
        if Workbook is None:
            raise RuntimeError("openpyxl required for .xlsx reports")
        wb = Workbook()
        ws = wb.active
        ws.append(fieldnames)
        for row in rows:
            ws.append([row.get(k, "") for k in fieldnames])
        wb.save(p)
    else:
        with p.open("w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f"  Report written: {p}")


# ══════════════════════════════════════════════════════════════════════════════
# §7  ENRICHERS
# ══════════════════════════════════════════════════════════════════════════════

class _BaseEnricher:
    def __init__(self, supabase: Client, *, apply: bool, limit: int = 0):
        self.sb = supabase
        self.apply = apply
        self.limit = limit or None
        self._counter = [-1]

    def _ensure_tag(self, name: str, tags_map: dict[str, int]) -> Optional[int]:
        return ensure_tag(self.sb, name, tags_map, apply=self.apply,
                          _synthetic_counter=self._counter)

    def _load_places(self, select: str) -> list[dict]:
        rows = fetch_all_rows(self.sb, "places", select)
        return rows[: self.limit] if self.limit else rows

    def _apply_links(self, links: list[dict]) -> int:
        return insert_links_batch(self.sb, links, apply=self.apply)


# ── 7a  Location enricher ─────────────────────────────────────────────────────

class LocationEnricher(_BaseEnricher):
    """
    Tags every place with its planning area, road name, and nearby MRT stations.
    """

    def __init__(
        self,
        supabase: Client,
        *,
        apply: bool,
        limit: int = 0,
        mrt_radius: float = 600.0,
        use_onemap: bool = False,
        skip_mrt: bool = False,
        skip_area: bool = False,
        onemap_token: Optional[str] = None,
    ):
        super().__init__(supabase, apply=apply, limit=limit)
        self.mrt_radius = mrt_radius
        self.use_onemap = use_onemap
        self.skip_mrt = skip_mrt
        self.skip_area = skip_area
        self._onemap_token = onemap_token or os.environ.get("ONEMAP_TOKEN", "")

    def _onemap_area(self, lat: float, lng: float) -> Optional[str]:
        try:
            params = urllib.parse.urlencode({"location": f"{lat},{lng}"})
            url = f"https://www.onemap.gov.sg/api/public/revgeocodexy?{params}"
            req = urllib.request.Request(url, headers={"Authorization": self._onemap_token})
            with urllib.request.urlopen(req, timeout=10) as resp:
                payload = json.loads(resp.read().decode())
            for info in payload.get("GeocodeInfo") or []:
                area = get_planning_area_from_postal(str(info.get("POSTALCODE") or ""))
                if area:
                    return area
        except Exception as e:
            print(f"    [warn] OneMap failed ({lat},{lng}): {e}")
        return None

    def run(self, *, report: str = "") -> dict:
        places = self._load_places("id, name, address, latitude, longitude")
        tags_map = fetch_tags_map(self.sb)
        existing = fetch_existing_links(self.sb, [p["id"] for p in places if p.get("id")])

        links: list[dict] = []
        report_rows: list[dict] = []
        new_links_total = 0

        for idx, place in enumerate(places, 1):
            pid, name = place.get("id"), place.get("name", "")
            addr = place.get("address", "")
            lat, lng = to_float(place.get("latitude")), to_float(place.get("longitude"))
            proposed: list[str] = []

            if not self.skip_area:
                area = (
                    self._onemap_area(lat, lng) if (self.use_onemap and lat and lng) else None
                ) or area_from_address(addr)
                if area:
                    proposed.append(area)
                    if self.use_onemap and lat:
                        time.sleep(0.15)
                # Road name tag — more specific than planning area
                road = extract_road_name_from_address(addr)
                if road:
                    proposed.append(road)

            if not self.skip_mrt and lat and lng:
                proposed.extend(nearby_mrt_tags(lat, lng, self.mrt_radius))

            # deduplicate case-insensitively
            seen: set[str] = set()
            deduped = [t for t in proposed if not (t.lower() in seen or seen.add(t.lower()))]

            new_for_place = 0
            for tag_name in deduped:
                tid = self._ensure_tag(tag_name, tags_map)
                if tid is None or pid is None:
                    continue
                pair = (pid, tid)
                if pair not in existing:
                    existing.add(pair)
                    links.append({"place_id": pid, "tag_id": tid})
                    new_links_total += 1
                    new_for_place += 1

            tag_str = ", ".join(deduped[:5]) + ("…" if len(deduped) > 5 else "")
            print(f"  [{idx}/{len(places)}] {name[:40]:<40} | tags: {tag_str} | new: {new_for_place}")
            report_rows.append({"place_id": pid, "place_name": name,
                                 "tags": "|".join(deduped), "new_links": new_for_place})

        inserted = self._apply_links(links)
        if report:
            write_report(report, report_rows, ["place_id", "place_name", "tags", "new_links"])
        return {"places": len(places), "new_links": new_links_total, "inserted": inserted}


# ── 7b  Google Places enricher ────────────────────────────────────────────────

class GoogleEnricher(_BaseEnricher):
    """
    Enriches places by calling Google Places Details API.
    Now also creates road-name tags and expands sub-cuisines to parent cuisines.
    """

    LEGACY_FIELDS = ("place_id,name,formatted_address,types,price_level,"
                     "reviews,editorial_summary")
    NEW_FIELDS = ("addressComponents,formattedAddress,shortFormattedAddress,"
                  "postalAddress,types,primaryType,displayName,priceLevel,"
                  "editorialSummary,reviews")

    def __init__(
        self,
        supabase: Client,
        *,
        apply: bool,
        limit: int = 0,
        api_key: str = "",
        use_find: bool = False,
        find_radius: int = 1000,
        include_allergies: bool = False,
        places_new: bool = False,
        fields: str = "",
        skip_if_tagged: bool = False,
        sleep_sec: float = 0.05,
    ):
        super().__init__(supabase, apply=apply, limit=limit)
        self.api_key = api_key or os.environ.get("GOOGLE_PLACES_API_KEY") or \
                       os.environ.get("GOOGLE_MAPS_API_KEY", "")
        self.use_find = use_find
        self.find_radius = find_radius
        self.include_allergies = include_allergies
        self.places_new = places_new
        self.fields = fields or (self.NEW_FIELDS if places_new else self.LEGACY_FIELDS)
        self.skip_if_tagged = skip_if_tagged
        self.sleep_sec = sleep_sec

    def _extract_place_id(self, uri: str) -> Optional[str]:
        if not uri:
            return None
        try:
            parsed = urllib.parse.urlparse(uri)
            params = urllib.parse.parse_qs(parsed.query)
            for key in ("place_id", "placeid"):
                if key in params:
                    return params[key][0]
            for q in params.get("q", []):
                if q.startswith("place_id:"):
                    return q.split("place_id:", 1)[1]
        except Exception:
            pass
        return None

    def _get_details(self, place_id: str) -> Optional[dict]:
        if self.places_new:
            url = f"https://places.googleapis.com/v1/places/{urllib.parse.quote(place_id)}"
            headers = {"X-Goog-Api-Key": self.api_key}
            if self.fields:
                headers["X-Goog-FieldMask"] = self.fields
            try:
                with urllib.request.urlopen(
                    urllib.request.Request(url, headers=headers), timeout=10
                ) as resp:
                    data = json.loads(resp.read().decode())
                    return None if "error" in data else data
            except Exception:
                return None
        else:
            params = urllib.parse.urlencode({
                "place_id": place_id, "fields": self.fields, "key": self.api_key,
            })
            url = f"https://maps.googleapis.com/maps/api/place/details/json?{params}"
            try:
                with urllib.request.urlopen(url, timeout=10) as resp:
                    data = json.loads(resp.read().decode())
                    return data.get("result") if data.get("status") == "OK" else None
            except Exception:
                return None
            finally:
                time.sleep(self.sleep_sec)

    def _find_place_id(self, query: str, lat: Optional[float], lng: Optional[float]) -> Optional[str]:
        params: dict[str, str] = {
            "input": query, "inputtype": "textquery",
            "fields": "place_id", "key": self.api_key,
        }
        if lat and lng:
            params["locationbias"] = f"circle:{self.find_radius}@{lat},{lng}"
        url = ("https://maps.googleapis.com/maps/api/place/findplacefromtext/json?"
               + urllib.parse.urlencode(params))
        try:
            with urllib.request.urlopen(url, timeout=10) as resp:
                data = json.loads(resp.read().decode())
                if data.get("status") == "OK":
                    cands = data.get("candidates") or []
                    return cands[0].get("place_id") if cands else None
        except Exception:
            pass
        finally:
            time.sleep(self.sleep_sec)
        return None

    def run(self, *, report: str = "") -> dict:
        places = self._load_places("id, gmaps_place_id, gmaps_uri, name, address, latitude, longitude")
        tags_map = fetch_tags_map(self.sb)
        tagged_ids: set[str] = set()
        if self.skip_if_tagged:
            tagged_ids = {str(r["place_id"])
                         for r in fetch_all_rows(self.sb, "place_tags", "place_id")}
        existing = fetch_existing_links(self.sb, [p["id"] for p in places if p.get("id")])

        links: list[dict] = []
        report_rows: list[dict] = []
        skipped_no_id = skipped_tagged = 0
        new_links_total = 0

        for idx, place in enumerate(places, 1):
            pid = str(place.get("id", ""))
            name = place.get("name", "")
            print(f"  [{idx}/{len(places)}] {name}")

            if self.skip_if_tagged and pid in tagged_ids:
                skipped_tagged += 1
                continue

            gid = place.get("gmaps_place_id") or self._extract_place_id(str(place.get("gmaps_uri") or ""))
            if not gid and self.use_find:
                q = " ".join(p for p in [name, place.get("address", "")] if p)
                if q:
                    gid = self._find_place_id(q, to_float(place.get("latitude")),
                                              to_float(place.get("longitude")))
            if not gid:
                skipped_no_id += 1
                report_rows.append({"place_id": pid, "place_name": name,
                                    "status": "missing_place_id", "tags": ""})
                continue

            details = self._get_details(gid)
            if not details:
                report_rows.append({"place_id": pid, "place_name": name,
                                    "status": "details_not_found", "tags": ""})
                continue

            reviews_text = extract_reviews_text(details.get("reviews"))
            editorial = ""
            for key in ("editorial_summary", "editorialSummary"):
                raw = details.get(key)
                editorial = (raw.get("overview") or raw.get("text") or ""
                             if isinstance(raw, dict) else str(raw or "")).strip()
                if editorial:
                    break
            details_name = extract_google_display_name(details) or name
            types = extract_google_types(details)
            address_text = extract_address_text_from_details(details) or str(place.get("address") or "")
            cuisine_text = " ".join(filter(None, [details_name, " ".join(types), editorial, address_text]))
            combined = " ".join(filter(None, [cuisine_text, reviews_text]))

            type_cuisines = infer_cuisine_tags_from_types(types)   # already expands parents
            text_cuisines = infer_cuisine_tags(cuisine_text, details_name)        # already expands parents
            cuisine_tags = type_cuisines + [t for t in text_cuisines if t not in type_cuisines]
            cuisine_tags = filter_cuisine_tags_for_venue(types, cuisine_tags, place_name=details_name)
            attr_tags = infer_attribute_tags_from_types(types)
            allergy_tags = infer_allergy_tags(combined) if self.include_allergies else []
            price_tag = infer_price_range_tag(
                details.get("price_level") or details.get("priceLevel"), combined
            )
            area_tag = infer_area_tag_from_details(details)
            # Road-name tag from the Supabase address column (most reliable source)
            road_tag = extract_road_name_from_address(place.get("address") or "") or extract_road_name_from_address(address_text)
            # Location tags inferred from the restaurant name itself
            name_location_tags = infer_location_tags_from_name(details_name)

            proposed_names = [t for t in [area_tag] + ([road_tag] if road_tag else []) + name_location_tags + attr_tags + cuisine_tags + [price_tag] + allergy_tags if t]

            # deduplicate
            seen: set[str] = set()
            deduped = [t for t in proposed_names
                       if is_english_tag(t) and not (t.lower() in seen or seen.add(t.lower()))]

            row_links = 0
            for tag_name in deduped:
                tid = self._ensure_tag(tag_name, tags_map)
                if tid is None:
                    continue
                pair_key = (int(pid), tid)
                if pair_key not in existing:
                    existing.add(pair_key)
                    links.append({"place_id": int(pid), "tag_id": tid})
                    new_links_total += 1
                    row_links += 1

            report_rows.append({"place_id": pid, "place_name": name, "status": "ok",
                                 "tags": "|".join(deduped), "new_links": row_links})
            print(f"    → {', '.join(deduped[:6])}")

        inserted = self._apply_links(links)
        if report:
            write_report(report, report_rows,
                         ["place_id", "place_name", "status", "tags", "new_links"])
        return {
            "places": len(places), "new_links": new_links_total, "inserted": inserted,
            "skipped_no_id": skipped_no_id, "skipped_tagged": skipped_tagged,
        }


# ── 7c  Dataset / Supabase auto-tagger ────────────────────────────────────────

@dataclass
class _DataRow:
    row_num: int
    values: dict[str, Any]
    _source: str = "supabase"

    @property
    def place_name(self) -> str:
        return str(self.values.get("name") or self.values.get("place_name") or "").strip()

    @property
    def address(self) -> str:
        return str(self.values.get("address") or self.values.get("formatted_address") or "").strip()

    @property
    def editorial_summary(self) -> str:
        return str(self.values.get("editorial_summary") or "").strip()

    @property
    def label_name(self) -> str:
        return str(self.values.get("label_name") or "").strip()

    @property
    def gmaps_place_id(self) -> str:
        return str(self.values.get("gmaps_place_id") or "").strip()

    @property
    def gmaps_uri(self) -> str:
        return str(self.values.get("gmaps_uri") or "").strip()

    @property
    def latitude(self) -> Optional[float]:
        return to_float(self.values.get("latitude"))

    @property
    def longitude(self) -> Optional[float]:
        return to_float(self.values.get("longitude"))


class AutoTagger(_BaseEnricher):
    """
    Infers area, road name, cuisine (with parent expansion), budget, and allergy
    tags from existing DB fields or an Excel dataset, then writes them to Supabase.
    """

    def __init__(
        self,
        supabase: Client,
        *,
        apply: bool,
        limit: int = 0,
        source: str = "supabase",
        dataset_path: str = "",
        sheet_name: str = "Result 1",
        use_google_geocode: bool = False,
        google_api_key: str = "",
        include_allergies: bool = False,
        skip_if_tagged: bool = False,
    ):
        super().__init__(supabase, apply=apply, limit=limit)
        self.source = source
        self.dataset_path = dataset_path
        self.sheet_name = sheet_name
        self.use_google_geocode = use_google_geocode
        self.google_api_key = (google_api_key
                               or os.environ.get("GOOGLE_PLACES_API_KEY")
                               or os.environ.get("GOOGLE_MAPS_API_KEY", ""))
        self.include_allergies = include_allergies
        self.skip_if_tagged = skip_if_tagged

    def _reverse_geocode_area(self, lat: float, lng: float) -> Optional[str]:
        if not self.google_api_key:
            return None
        params = urllib.parse.urlencode({
            "latlng": f"{lat},{lng}", "key": self.google_api_key, "language": "en",
        })
        url = f"https://maps.googleapis.com/maps/api/geocode/json?{params}"
        try:
            with urllib.request.urlopen(url, timeout=8) as resp:
                payload = json.loads(resp.read().decode())
            if payload.get("status") == "OK":
                for result in payload.get("results", []):
                    area = _extract_area_from_address_components(
                        result.get("address_components", []))
                    if area:
                        return area
        except Exception as e:
            print(f"    [warn] Geocode failed ({lat},{lng}): {e}")
        finally:
            time.sleep(0.03)
        return None

    def _infer_area(self, row: _DataRow) -> Optional[str]:
        if self.use_google_geocode and row.latitude and row.longitude:
            area = self._reverse_geocode_area(row.latitude, row.longitude)
            if area:
                return sanitize_area_candidate(area)
        return area_from_address(row.address) or sanitize_area_candidate(
            str(self.values_get(row, "geography") or "")
        )

    @staticmethod
    def values_get(row: _DataRow, key: str) -> Any:
        return row.values.get(key)

    def _infer_tags_for_row(self, row: _DataRow) -> list[str]:
        combined = " ".join(filter(None, [
            row.place_name, row.label_name, row.editorial_summary,
            extract_reviews_text(row.values.get("reviews")),
        ]))
        area = self._infer_area(row)
        # Road-name tag — fine-grained location tag from address
        road = extract_road_name_from_address(row.address)
        price_tag = infer_price_range_tag(row.values.get("price_level"), combined)
        # infer_cuisine_tags already calls expand_with_parent_cuisines internally
        cuisine_tags = infer_cuisine_tags(combined, row.label_name)
        allergy_tags = infer_allergy_tags(combined) if self.include_allergies else []
        road_tags = [road] if road else []
        # Location tags inferred from the restaurant name itself
        name_location_tags = infer_location_tags_from_name(row.place_name)
        proposed = [t for t in ([area] + road_tags + [price_tag] + cuisine_tags + allergy_tags + name_location_tags) if t]
        seen: set[str] = set()
        return [t for t in proposed
                if is_english_tag(t) and not (t.lower() in seen or seen.add(t.lower()))]

    def _load_dataset_rows(self) -> list[_DataRow]:
        if load_workbook is None:
            raise RuntimeError("openpyxl required for dataset source")
        wb = load_workbook(self.dataset_path, read_only=True, data_only=True)
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' not found in {self.dataset_path}")
        ws = wb[self.sheet_name]
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(it)]
        rows = [_DataRow(i, dict(zip(headers, vals)), "dataset")
                for i, vals in enumerate(it, 1)]
        return rows[: self.limit] if self.limit else rows

    def _load_supabase_rows(self) -> list[_DataRow]:
        places = fetch_all_rows(
            self.sb, "places",
            "id, name, address, gmaps_place_id, gmaps_uri, "
            "editorial_summary, latitude, longitude, price_level"
        )
        if self.limit:
            places = places[: self.limit]
        return [_DataRow(i, p, "supabase") for i, p in enumerate(places, 1)]

    def _match_place_id(self, row: _DataRow, place_lookup: dict) -> Optional[str]:
        raw_id = row.values.get("id")
        if raw_id is not None and str(raw_id) in place_lookup:
            return str(raw_id)
        gid = row.gmaps_place_id
        if gid and gid in place_lookup.get("_by_gmaps_id", {}):
            return place_lookup["_by_gmaps_id"][gid]
        key = f"{normalize_text(row.place_name)}||{normalize_text(row.address)}"
        return place_lookup.get("_by_name_addr", {}).get(key)

    def run(self, *, report: str = "") -> dict:
        if self.source == "dataset":
            rows = self._load_dataset_rows()
        else:
            rows = self._load_supabase_rows()

        place_lookup: dict[str, Any] = {}
        if self.source == "dataset":
            all_places = fetch_all_rows(
                self.sb, "places",
                "id, name, address, gmaps_place_id, gmaps_uri"
            )
            place_lookup["_by_gmaps_id"] = {
                p["gmaps_place_id"]: str(p["id"])
                for p in all_places if p.get("gmaps_place_id")
            }
            place_lookup["_by_name_addr"] = {
                f"{normalize_text(p.get('name',''))}||{normalize_text(p.get('address',''))}": str(p["id"])
                for p in all_places
            }
            place_lookup.update({str(p["id"]): str(p["id"]) for p in all_places})

        tags_map = fetch_tags_map(self.sb)
        all_place_ids = [
            int(r.values.get("id")) for r in rows
            if r.values.get("id") is not None and str(r.values.get("id")).lstrip("-").isdigit()
        ]
        existing = fetch_existing_links(self.sb, all_place_ids)
        tagged_ids: set[str] = set()
        if self.skip_if_tagged:
            tagged_ids = {str(rec["place_id"])
                         for rec in fetch_all_rows(self.sb, "place_tags", "place_id")}

        links: list[dict] = []
        report_rows: list[dict] = []
        unmatched = new_links_total = 0

        for row in rows:
            if self.source == "dataset":
                pid_str = self._match_place_id(row, place_lookup)
                if not pid_str:
                    unmatched += 1
                    report_rows.append({"row": row.row_num, "place": row.place_name,
                                        "matched": False, "tags": "", "new_links": 0})
                    continue
            else:
                pid_str = str(row.values.get("id", ""))

            if self.skip_if_tagged and pid_str in tagged_ids:
                continue

            try:
                pid_int = int(pid_str)
            except (ValueError, TypeError):
                unmatched += 1
                continue

            tag_names = self._infer_tags_for_row(row)
            row_links = 0
            for tag_name in tag_names:
                tid = self._ensure_tag(tag_name, tags_map)
                if tid is None:
                    continue
                pair = (pid_int, tid)
                if pair not in existing:
                    existing.add(pair)
                    links.append({"place_id": pid_int, "tag_id": tid})
                    new_links_total += 1
                    row_links += 1

            report_rows.append({"row": row.row_num, "place": row.place_name,
                                 "matched": True, "tags": "|".join(tag_names),
                                 "new_links": row_links})
            print(f"  [{row.row_num}] {row.place_name[:40]:<40} | {', '.join(tag_names[:5])}")

        inserted = self._apply_links(links)
        if report:
            write_report(report, report_rows,
                         ["row", "place", "matched", "tags", "new_links"])
        return {
            "rows": len(rows), "unmatched": unmatched,
            "new_links": new_links_total, "inserted": inserted,
        }


# ══════════════════════════════════════════════════════════════════════════════
# §8  CLI
# ══════════════════════════════════════════════════════════════════════════════

def _build_parser() -> argparse.ArgumentParser:
    root = argparse.ArgumentParser(
        prog="enrich",
        description="FoodKakiBot — unified tagging & location enrichment",
    )
    sub = root.add_subparsers(dest="cmd", required=True)

    common = argparse.ArgumentParser(add_help=False)
    common.add_argument("--apply", action="store_true")
    common.add_argument("--limit", type=int, default=0)
    common.add_argument("--report", default="")
    common.add_argument("--skip-if-tagged", action="store_true")

    p_loc = sub.add_parser("location", parents=[common])
    p_loc.add_argument("--mrt-radius", type=int, default=600)
    p_loc.add_argument("--use-onemap", action="store_true")
    p_loc.add_argument("--onemap-token", default="")
    p_loc.add_argument("--skip-mrt", action="store_true")
    p_loc.add_argument("--skip-area", action="store_true")

    p_goog = sub.add_parser("google", parents=[common])
    p_goog.add_argument("--api-key", default="")
    p_goog.add_argument("--use-find", action="store_true")
    p_goog.add_argument("--find-radius", type=int, default=1000)
    p_goog.add_argument("--include-allergies", action="store_true")
    p_goog.add_argument("--places-new", action="store_true")
    p_goog.add_argument("--fields", default="")
    p_goog.add_argument("--sleep", type=float, default=0.05)

    p_auto = sub.add_parser("auto", parents=[common])
    p_auto.add_argument("--source", choices=["supabase", "dataset"], default="supabase")
    p_auto.add_argument("--dataset", default="")
    p_auto.add_argument("--sheet", default="Result 1")
    p_auto.add_argument("--use-google-geocode", action="store_true")
    p_auto.add_argument("--google-api-key", default="")
    p_auto.add_argument("--include-allergies", action="store_true")

    p_all = sub.add_parser("all", parents=[common])
    p_all.add_argument("--mrt-radius", type=int, default=600)
    p_all.add_argument("--use-onemap", action="store_true")
    p_all.add_argument("--skip-mrt", action="store_true")
    p_all.add_argument("--skip-area", action="store_true")
    p_all.add_argument("--use-find", action="store_true")
    p_all.add_argument("--include-allergies", action="store_true")
    p_all.add_argument("--places-new", action="store_true")
    p_all.add_argument("--use-google-geocode", action="store_true")

    return root


def _make_supabase() -> Client:
    if create_client is None:
        raise RuntimeError("supabase-py not installed — run: pip install supabase")
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_ANON_KEY")
    if not url or not key:
        raise RuntimeError("Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY in your .env")
    return create_client(url, key)


def _print_stats(label: str, stats: dict) -> None:
    print(f"\n── {label} ────────────────────────────────────")
    for k, v in stats.items():
        print(f"   {k:<20}: {v}")


def main(argv: Optional[list[str]] = None) -> None:
    load_dotenv()
    args = _build_parser().parse_args(argv)
    sb = _make_supabase()

    mode = "APPLY" if args.apply else "DRY-RUN"
    print(f"\nFoodKakiBot Enricher  [{mode}]  cmd={args.cmd}\n{'='*55}")

    if args.cmd == "location":
        enricher = LocationEnricher(
            sb, apply=args.apply, limit=args.limit,
            mrt_radius=args.mrt_radius, use_onemap=args.use_onemap,
            skip_mrt=args.skip_mrt, skip_area=args.skip_area,
            onemap_token=args.onemap_token,
        )
        _print_stats("location", enricher.run(report=args.report))

    elif args.cmd == "google":
        enricher = GoogleEnricher(
            sb, apply=args.apply, limit=args.limit,
            api_key=args.api_key, use_find=args.use_find,
            find_radius=args.find_radius,
            include_allergies=args.include_allergies,
            places_new=args.places_new, fields=args.fields,
            skip_if_tagged=args.skip_if_tagged, sleep_sec=args.sleep,
        )
        _print_stats("google", enricher.run(report=args.report))

    elif args.cmd == "auto":
        if args.source == "dataset" and not args.dataset:
            sys.exit("--dataset PATH is required when --source=dataset")
        enricher = AutoTagger(
            sb, apply=args.apply, limit=args.limit,
            source=args.source, dataset_path=args.dataset,
            sheet_name=args.sheet, use_google_geocode=args.use_google_geocode,
            google_api_key=args.google_api_key,
            include_allergies=args.include_allergies,
            skip_if_tagged=args.skip_if_tagged,
        )
        _print_stats("auto", enricher.run(report=args.report))

    elif args.cmd == "all":
        common = dict(apply=args.apply, limit=args.limit)
        for label, enricher in [
            ("location", LocationEnricher(
                sb, **common, mrt_radius=args.mrt_radius,
                use_onemap=args.use_onemap, skip_mrt=args.skip_mrt,
                skip_area=args.skip_area,
            )),
            ("auto", AutoTagger(
                sb, **common, use_google_geocode=args.use_google_geocode,
                include_allergies=args.include_allergies,
            )),
            ("google", GoogleEnricher(
                sb, **common, use_find=args.use_find,
                include_allergies=args.include_allergies,
                places_new=args.places_new,
                skip_if_tagged=args.skip_if_tagged,
            )),
        ]:
            print(f"\n{'─'*40}\nRunning: {label}\n{'─'*40}")
            _print_stats(label, enricher.run())

    if not args.apply:
        print("\n  (Dry-run — pass --apply to write changes to Supabase)")


if __name__ == "__main__":
    main()
