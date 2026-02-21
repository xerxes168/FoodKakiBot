#!/usr/bin/env python3
"""Auto-tag places in Supabase using dataset or places-table signals.

What this script does:
1. Loads source rows from Supabase `places` (default) or an Excel dataset.
2. Infers area tags (lat/lng reverse geocode when available).
3. Infers cuisine, budget, and allergy tags from reviews/metadata.
4. Reuses existing tags and creates missing tags in `tags`.
5. Creates missing links in `place_tags`.

Run dry-run first, then apply if output looks correct.
"""

from __future__ import annotations

import argparse
import ast
import csv
import json
import os
import re
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
except Exception:
    def load_dotenv() -> bool:  # type: ignore[misc]
        return False

try:
    from supabase import Client, create_client
except Exception:
    Client = Any  # type: ignore[misc,assignment]
    create_client = None  # type: ignore[assignment]


CUISINE_KEYWORDS: Dict[str, Sequence[str]] = {
    "Japanese": (
        "japanese",
        "sushi",
        "sashimi",
        "ramen",
        "udon",
        "soba",
        "yakitori",
        "izakaya",
        "tempura",
        "tonkatsu",
        "donburi",
    ),
    "Korean": (
        "korean",
        "kimchi",
        "bibimbap",
        "bulgogi",
        "tteokbokki",
        "jjajangmyeon",
        "banchan",
        "samgyeopsal",
    ),
    "Chinese": (
        "chinese",
        "dim sum",
        "dumpling",
        "szechuan",
        "sichuan",
        "cantonese",
        "wonton",
        "xiao long bao",
        "char siew",
    ),
    "Mala": ("mala", "ma la", "spicy pot"),
    "Indian": (
        "indian",
        "biryani",
        "tandoori",
        "naan",
        "masala",
        "paneer",
        "roti prata",
        "thosai",
    ),
    "Malay": (
        "malay",
        "nasi lemak",
        "satay",
        "rendang",
        "mee rebus",
        "lontong",
    ),
    "Thai": ("thai", "tom yum", "green curry", "pad thai", "som tam"),
    "Vietnamese": ("vietnamese", "pho", "banh mi", "bun cha", "spring roll"),
    "Italian": ("italian", "pasta", "risotto", "pizza", "lasagna", "carbonara"),
    "French": ("french", "confit", "croissant", "escargot", "foie gras"),
    "Mexican": ("mexican", "taco", "burrito", "quesadilla", "nachos"),
    "Western": ("western", "steak", "burger", "fish and chips", "brunch"),
    "Seafood": ("seafood", "crab", "lobster", "oyster", "prawn", "clam"),
    "Vegetarian": ("vegetarian", "vegan", "plant-based", "meat-free"),
    "Halal": ("halal",),
    "Dessert": ("dessert", "cake", "pastry", "gelato", "ice cream", "sweet"),
    "Cafe": ("cafe", "coffee", "latte", "espresso", "brunch"),
    "Bubble Tea": ("bubble tea", "boba", "milk tea", "pearls"),
}

CUISINE_ALIASES: Dict[str, str] = {
    "jpn": "Japanese",
    "jp": "Japanese",
    "korea": "Korean",
    "chinese food": "Chinese",
    "veg": "Vegetarian",
    "western food": "Western",
}

ALLERGY_KEYWORDS: Dict[str, Sequence[str]] = {
    "Gluten-Free": ("gluten free", "gluten-free", "gf"),
    "Dairy-Free": ("dairy free", "dairy-free", "lactose free", "lactose-free", "no dairy"),
    "Nut-Free": ("nut free", "nut-free", "peanut-free", "no nuts", "tree nut"),
    "Shellfish-Free": ("shellfish free", "shellfish-free", "no shellfish"),
    "Egg-Free": ("egg free", "egg-free", "no egg"),
    "Soy-Free": ("soy free", "soy-free", "no soy"),
}

AREA_COMPONENT_TYPES = (
    "neighborhood",
    "sublocality_level_1",
    "sublocality",
    "locality",
    "administrative_area_level_2",
)

POINT_WKT_RE = re.compile(r"^\s*point\s*\(.+\)\s*$", re.IGNORECASE)
NON_ASCII_RE = re.compile(r"[^\x00-\x7F]")
PRICE_NUMBER_RE = re.compile(r"\$(\d+(?:\.\d{1,2})?)")

PRICE_TAG_BUDGET = "Budget"
PRICE_TAG_MID_RANGE = "Mid Range"
PRICE_TAG_EXPENSIVE = "Expensive"
UNKNOWN_AREA_TAG = "Unknown Area"
TAG_CATEGORY_BUDGET = "budget"
TAG_CATEGORY_CUISINE = "cuisine"
TAG_CATEGORY_ALLERGY = "allergy"
TAG_CATEGORY_AREA = "area"


@dataclass
class DatasetRow:
    row_num: int
    values: Dict[str, Any]

    @property
    def place_name(self) -> str:
        return str(self.values.get("place_name") or "").strip()

    @property
    def formatted_address(self) -> str:
        return str(self.values.get("formatted_address") or "").strip()

    @property
    def gmaps_uri(self) -> str:
        return str(self.values.get("gmaps_uri") or "").strip()

    @property
    def gmaps_place_id(self) -> str:
        return str(self.values.get("gmaps_place_id") or "").strip()

    @property
    def geography(self) -> str:
        return str(self.values.get("geography") or "").strip()

    @property
    def label_name(self) -> str:
        return str(self.values.get("label_name") or "").strip()

    @property
    def editorial_summary(self) -> str:
        return str(self.values.get("editorial_summary") or "").strip()

    @property
    def latitude(self) -> Optional[float]:
        return to_float(self.values.get("latitude"))

    @property
    def longitude(self) -> Optional[float]:
        return to_float(self.values.get("longitude"))


@dataclass
class PlaceRow:
    row_num: int
    values: Dict[str, Any]

    @property
    def place_name(self) -> str:
        return str(self.values.get("name") or "").strip()

    @property
    def formatted_address(self) -> str:
        return str(self.values.get("address") or "").strip()

    @property
    def gmaps_uri(self) -> str:
        return str(self.values.get("gmaps_uri") or "").strip()

    @property
    def gmaps_place_id(self) -> str:
        return str(self.values.get("gmaps_place_id") or "").strip()

    @property
    def geography(self) -> str:
        return str(self.values.get("location") or "").strip()

    @property
    def label_name(self) -> str:
        return ""

    @property
    def editorial_summary(self) -> str:
        return str(self.values.get("editorial_summary") or "").strip()

    @property
    def latitude(self) -> Optional[float]:
        return to_float(self.values.get("latitude"))

    @property
    def longitude(self) -> Optional[float]:
        return to_float(self.values.get("longitude"))


class GoogleReverseGeocoder:
    def __init__(self, api_key: str, delay_sec: float = 0.03) -> None:
        self.api_key = api_key
        self.delay_sec = delay_sec

    def reverse_area(self, lat: float, lng: float) -> Optional[str]:
        params = urllib.parse.urlencode(
            {
                "latlng": f"{lat},{lng}",
                "key": self.api_key,
                "language": "en",
            }
        )
        url = f"https://maps.googleapis.com/maps/api/geocode/json?{params}"
        try:
            with urllib.request.urlopen(url, timeout=8) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
        except Exception:
            return None
        finally:
            if self.delay_sec > 0:
                time.sleep(self.delay_sec)

        if payload.get("status") != "OK":
            return None

        for result in payload.get("results", []):
            for comp in result.get("address_components", []):
                types = set(comp.get("types", []))
                for area_type in AREA_COMPONENT_TYPES:
                    if area_type in types:
                        name = str(comp.get("long_name") or "").strip()
                        if name:
                            return name
        return None


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", " ", text)


def normalize_tag_name(value: str) -> str:
    text = re.sub(r"\s+", " ", (value or "").strip())
    return text


def is_english_tag(value: str) -> bool:
    candidate = normalize_tag_name(value)
    if not candidate:
        return False
    if NON_ASCII_RE.search(candidate):
        return False
    return bool(re.search(r"[A-Za-z]", candidate))


def sanitize_area_candidate(value: str) -> Optional[str]:
    candidate = normalize_tag_name(value)
    if not candidate:
        return None
    if not is_english_tag(candidate):
        return None
    if POINT_WKT_RE.match(candidate):
        return None
    if candidate.lower() in {"singapore", "sg"}:
        return None
    if re.search(r"\d", candidate):
        return None
    if "," in candidate:
        return None
    if len(candidate) > 40:
        return None
    return candidate


def to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def safe_json_loads(raw: Any) -> Any:
    if raw is None:
        return None
    if isinstance(raw, (dict, list)):
        return raw
    text = str(raw).strip()
    if not text:
        return None

    try:
        return json.loads(text)
    except Exception:
        pass

    try:
        return ast.literal_eval(text)
    except Exception:
        return None


def extract_reviews_text(raw_reviews: Any) -> str:
    parsed = safe_json_loads(raw_reviews)
    if not parsed:
        return ""

    texts: List[str] = []

    if isinstance(parsed, dict) and "reviews" in parsed:
        parsed = parsed["reviews"]

    if isinstance(parsed, list):
        for item in parsed:
            if isinstance(item, dict):
                review_text = str(item.get("text") or "").strip()
                if review_text:
                    texts.append(review_text)
            elif isinstance(item, str):
                if item.strip():
                    texts.append(item.strip())

    return "\n".join(texts)


def extract_area_from_gmaps_response(raw_response: Any) -> Optional[str]:
    payload = safe_json_loads(raw_response)
    if not isinstance(payload, dict):
        return None

    address_components = payload.get("address_components")
    if not isinstance(address_components, list):
        return None

    for preferred_type in AREA_COMPONENT_TYPES:
        for comp in address_components:
            types = set(comp.get("types") or [])
            if preferred_type in types:
                candidate = str(comp.get("long_name") or "").strip()
                if candidate:
                    return candidate

    return None


def extract_price_level_from_gmaps_response(raw_response: Any) -> Optional[int]:
    payload = safe_json_loads(raw_response)
    if not isinstance(payload, dict):
        return None
    level = payload.get("price_level")
    try:
        return int(level) if level is not None else None
    except (TypeError, ValueError):
        return None


def canonical_cuisine_name(name: str) -> str:
    key = normalize_text(name)
    if key in CUISINE_ALIASES:
        return CUISINE_ALIASES[key]

    for cuisine in CUISINE_KEYWORDS:
        if normalize_text(cuisine) == key:
            return cuisine

    return name.strip()


def infer_cuisine_tags(row: DatasetRow) -> List[str]:
    score: Dict[str, int] = {}

    combined_text_parts = [
        row.place_name,
        row.label_name,
        row.editorial_summary,
        extract_reviews_text(row.values.get("reviews")),
    ]
    text = normalize_text("\n".join(part for part in combined_text_parts if part))

    label = canonical_cuisine_name(row.label_name)
    if label in CUISINE_KEYWORDS:
        score[label] = score.get(label, 0) + 4

    for cuisine, keywords in CUISINE_KEYWORDS.items():
        for kw in keywords:
            kw_norm = normalize_text(kw)
            if not kw_norm:
                continue
            if kw_norm in text:
                score[cuisine] = score.get(cuisine, 0) + (2 if " " in kw_norm else 1)

    ranked = sorted(score.items(), key=lambda pair: (-pair[1], pair[0]))

    selected: List[str] = []
    for cuisine, points in ranked:
        if points < 2:
            continue
        if is_english_tag(cuisine):
            selected.append(cuisine)
        if len(selected) >= 3:
            break

    return selected


def infer_allergy_tags(row: DatasetRow) -> List[str]:
    combined_text_parts = [
        row.place_name,
        row.label_name,
        row.editorial_summary,
        extract_reviews_text(row.values.get("reviews")),
    ]
    text = normalize_text("\n".join(part for part in combined_text_parts if part))
    matched: List[str] = []
    for tag_name, keywords in ALLERGY_KEYWORDS.items():
        for kw in keywords:
            if normalize_text(kw) in text:
                matched.append(tag_name)
                break
    return matched


def infer_area_tag(row: DatasetRow, geocoder: Optional[GoogleReverseGeocoder]) -> str:
    from_payload = extract_area_from_gmaps_response(row.values.get("gmaps_response"))
    clean = sanitize_area_candidate(from_payload or "")
    if clean:
        return clean

    if geocoder and row.latitude is not None and row.longitude is not None:
        from_reverse = geocoder.reverse_area(row.latitude, row.longitude)
        clean = sanitize_area_candidate(from_reverse or "")
        if clean:
            return clean

    clean = sanitize_area_candidate(row.geography)
    if clean:
        return clean

    return UNKNOWN_AREA_TAG


def infer_price_range_tag(row: DatasetRow) -> str:
    # 1) Prefer Google's structured price_level when available.
    level = extract_price_level_from_gmaps_response(row.values.get("gmaps_response"))
    if level is not None:
        if level <= 1:
            return PRICE_TAG_BUDGET
        if level == 2:
            return PRICE_TAG_MID_RANGE
        return PRICE_TAG_EXPENSIVE

    # 2) Fallback: infer from text signals in reviews/metadata.
    combined = " ".join(
        [
            row.place_name,
            row.label_name,
            row.editorial_summary,
            extract_reviews_text(row.values.get("reviews")),
        ]
    )
    text = normalize_text(combined)

    if any(token in text for token in ("cheap", "affordable", "budget", "value for money", "wallet-friendly")):
        return PRICE_TAG_BUDGET
    if any(token in text for token in ("expensive", "pricey", "premium", "high-end", "fine dining", "$$$")):
        return PRICE_TAG_EXPENSIVE
    if "$$" in text:
        return PRICE_TAG_MID_RANGE

    amounts = [float(m.group(1)) for m in PRICE_NUMBER_RE.finditer(combined)]
    if amounts:
        avg = sum(amounts) / len(amounts)
        if avg <= 15:
            return PRICE_TAG_BUDGET
        if avg <= 35:
            return PRICE_TAG_MID_RANGE
        return PRICE_TAG_EXPENSIVE

    # 3) Mandatory fallback so every restaurant has one price-range tag.
    return PRICE_TAG_MID_RANGE


def infer_tag_category(tag_name: str, area_tag: str) -> Optional[str]:
    if not tag_name:
        return None
    if tag_name in (PRICE_TAG_BUDGET, PRICE_TAG_MID_RANGE, PRICE_TAG_EXPENSIVE):
        return TAG_CATEGORY_BUDGET
    if tag_name in CUISINE_KEYWORDS:
        return TAG_CATEGORY_CUISINE
    if tag_name in ALLERGY_KEYWORDS:
        return TAG_CATEGORY_ALLERGY
    if area_tag and tag_name == area_tag:
        return TAG_CATEGORY_AREA
    return None


def chunked(values: Sequence[Dict[str, Any]], size: int) -> Iterable[Sequence[Dict[str, Any]]]:
    for i in range(0, len(values), size):
        yield values[i : i + size]


def fetch_all_table_rows(supabase: Client, table: str, select_expr: str, page_size: int = 1000) -> List[Dict[str, Any]]:
    all_rows: List[Dict[str, Any]] = []
    offset = 0

    while True:
        response = (
            supabase.table(table)
            .select(select_expr)
            .range(offset, offset + page_size - 1)
            .execute()
        )
        rows = response.data or []
        all_rows.extend(rows)
        if len(rows) < page_size:
            break
        offset += page_size

    return all_rows


def pick_tags_query(supabase: Client) -> Tuple[str, List[Dict[str, Any]]]:
    candidates = [
        "id, name, category",
        "id, name",
    ]
    last_error: Optional[Exception] = None
    for select_expr in candidates:
        try:
            rows = fetch_all_table_rows(supabase, "tags", select_expr)
            return select_expr, rows
        except Exception as exc:
            last_error = exc
            continue
    raise RuntimeError(f"Unable to query tags table with supported columns. Last error: {last_error}")


def find_tag_by_name(supabase: Client, tag_name: str) -> Optional[Dict[str, Any]]:
    # Try exact first, then case-insensitive exact.
    lookups = [
        supabase.table("tags").select("id, name").eq("name", tag_name).limit(1),
        supabase.table("tags").select("id, name").ilike("name", tag_name).limit(1),
    ]
    for query in lookups:
        try:
            response = query.execute()
            row = (response.data or [None])[0]
            if row and row.get("id") is not None:
                return row
        except Exception:
            continue
    return None


def read_csv_rows(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader]
    return fieldnames, rows


def write_csv_rows(path: Path, fieldnames: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(fieldnames))
        writer.writeheader()
        writer.writerows(rows)


def load_dataset_rows(dataset_path: Path, sheet_name: str) -> Tuple[List[str], List[DatasetRow]]:
    wb = load_workbook(dataset_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    iterator = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(iterator)]

    rows: List[DatasetRow] = []
    for row_num, values in enumerate(iterator, start=1):
        row_dict = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        rows.append(DatasetRow(row_num=row_num, values=row_dict))

    return headers, rows


def pick_place_query(supabase: Client) -> Tuple[str, List[Dict[str, Any]]]:
    candidates = [
        "id, gmaps_place_id, gmaps_uri, name, address",
        "id, gmaps_uri, name, address",
        "id, name, address",
    ]

    last_error: Optional[Exception] = None
    for select_expr in candidates:
        try:
            rows = fetch_all_table_rows(supabase, "places", select_expr)
            return select_expr, rows
        except Exception as exc:
            last_error = exc
            continue

    raise RuntimeError(f"Unable to query places table with supported columns. Last error: {last_error}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Auto-tag restaurants into tags/place_tags using places or dataset signals")
    parser.add_argument(
        "--source",
        choices=("supabase", "dataset"),
        default="supabase",
        help="Source of rows to tag (default: supabase)",
    )
    parser.add_argument(
        "--dataset",
        default=str(Path(__file__).resolve().parents[2] / "datasets" / "food_places_data_set.xlsx"),
        help="Path to dataset xlsx (used when --source=dataset)",
    )
    parser.add_argument("--sheet", default="Result 1", help="Worksheet name in the dataset (used when --source=dataset)")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually write changes to Supabase (default is dry-run)",
    )
    parser.add_argument(
        "--use-google-geocode",
        action="store_true",
        help="Use Google reverse geocoding for area when fallback is needed",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional row limit for testing (0 means all rows)",
    )
    parser.add_argument(
        "--report",
        default="",
        help="Optional report output path (.csv, .json, or .xlsx). Example: ./tagging_report.xlsx",
    )
    parser.add_argument("--places-csv", default="", help="Path to exported places.csv")
    parser.add_argument("--tags-csv", default="", help="Path to exported tags.csv")
    parser.add_argument("--place-tags-csv", default="", help="Path to exported place_tags.csv")
    parser.add_argument(
        "--out-tags-csv",
        default="",
        help="Output CSV path for updated tags table (CSV mode)",
    )
    parser.add_argument(
        "--out-place-tags-csv",
        default="",
        help="Output CSV path for updated place_tags table (CSV mode)",
    )
    args = parser.parse_args()

    load_dotenv()

    csv_mode = bool(args.places_csv or args.tags_csv or args.place_tags_csv)
    if csv_mode and not (args.places_csv and args.tags_csv and args.place_tags_csv):
        raise RuntimeError("CSV mode requires --places-csv, --tags-csv, and --place-tags-csv.")

    dataset_path = Path(args.dataset).resolve()
    if args.source == "dataset" and not dataset_path.exists():
        raise FileNotFoundError(f"Dataset not found: {dataset_path}")

    geocoder: Optional[GoogleReverseGeocoder] = None
    if args.use_google_geocode:
        maps_key = os.environ.get("GOOGLE_PLACES_API_KEY") or os.environ.get("GOOGLE_MAPS_API_KEY")
        if maps_key:
            geocoder = GoogleReverseGeocoder(api_key=maps_key)
        else:
            print("[warn] --use-google-geocode set but no GOOGLE_MAPS_API_KEY/GOOGLE_API_KEY found; skipping reverse geocode")

    supabase: Optional[Client] = None
    if not csv_mode:
        if create_client is None:
            raise RuntimeError("Supabase SDK is not installed. Install requirements or use CSV mode.")
        supabase_url = os.environ.get("SUPABASE_URL")
        supabase_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_ANON_KEY")
        if not supabase_url or not supabase_key:
            raise RuntimeError("Missing SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY/SUPABASE_ANON_KEY")
        supabase = create_client(supabase_url, supabase_key)

    headers: List[str] = []
    dataset_rows: List[DatasetRow] = []
    if args.source == "dataset":
        headers, dataset_rows = load_dataset_rows(dataset_path, args.sheet)
        if args.limit > 0:
            dataset_rows = dataset_rows[: args.limit]

    print(f"Source: {args.source}")
    if args.source == "dataset":
        print(f"Dataset: {dataset_path}")
        print(f"Sheet: {args.sheet}")
        print(f"Rows loaded: {len(dataset_rows)}")
    if csv_mode:
        print("Mode: CSV-EXPORT")
    else:
        print(f"Mode: {'APPLY' if args.apply else 'DRY-RUN'}")

    if args.source == "dataset" and "id" not in headers:
        print("[info] Dataset has no 'id' column; matching will use gmaps_place_id/gmaps_uri/name+address.")

    tags_fieldnames: List[str]
    place_tags_fieldnames: List[str]
    tags_select = ""
    if csv_mode:
        places_csv_path = Path(args.places_csv).expanduser().resolve()
        tags_csv_path = Path(args.tags_csv).expanduser().resolve()
        place_tags_csv_path = Path(args.place_tags_csv).expanduser().resolve()
        _, place_rows = read_csv_rows(places_csv_path)
        tags_fieldnames, tags_rows = read_csv_rows(tags_csv_path)
        place_tags_fieldnames, place_tag_rows = read_csv_rows(place_tags_csv_path)
        place_select = "csv"
        tags_select = "csv"
    else:
        assert supabase is not None
        place_select, place_rows = pick_place_query(supabase)
        tags_select, tags_rows = pick_tags_query(supabase)
        place_tag_rows = fetch_all_table_rows(supabase, "place_tags", "place_id, tag_id")
        tags_fieldnames = ["id", "name"]
        if tags_select and "category" in tags_select:
            tags_fieldnames.append("category")
        place_tags_fieldnames = ["place_id", "tag_id"]

    if args.source == "supabase" and args.limit > 0:
        place_rows = place_rows[: args.limit]

    print(f"Places query used: {place_select}")
    print(f"Places loaded: {len(place_rows)}")
    if not csv_mode:
        print(f"Tags query used: {tags_select}")

    tags_by_norm: Dict[str, Dict[str, Any]] = {
        normalize_text(t.get("name")): t for t in tags_rows if t.get("name") is not None
    }
    tags_has_category = "category" in tags_fieldnames
    existing_pairs: Set[Tuple[str, str]] = {
        (str(r.get("place_id")), str(r.get("tag_id")))
        for r in place_tag_rows
        if r.get("place_id") is not None and r.get("tag_id") is not None
    }

    place_by_id = {str(p.get("id")): p for p in place_rows if p.get("id") is not None}
    place_by_gmaps_place_id = {
        normalize_text(p.get("gmaps_place_id")): p
        for p in place_rows
        if p.get("gmaps_place_id")
    }
    place_by_gmaps_uri = {
        normalize_text(p.get("gmaps_uri")): p
        for p in place_rows
        if p.get("gmaps_uri")
    }
    place_by_name_address = {
        f"{normalize_text(p.get('name'))}||{normalize_text(p.get('address'))}": p
        for p in place_rows
    }

    created_tags = 0
    planned_links = 0
    failed_link_inserts = 0
    failed_tag_ensures = 0
    unmatched_rows = 0
    rows_with_no_tags = 0
    place_tags_to_insert: List[Dict[str, Any]] = []
    merged_tags_rows: List[Dict[str, Any]] = [dict(r) for r in tags_rows]
    merged_place_tags_rows: List[Dict[str, Any]] = [dict(r) for r in place_tag_rows]
    report_rows: List[Dict[str, Any]] = []

    synthetic_tag_id = -1
    max_tag_id = 0
    for row in tags_rows:
        try:
            max_tag_id = max(max_tag_id, int(str(row.get("id"))))
        except Exception:
            continue

    def ensure_tag_id(tag_name: str, category: Optional[str]) -> Optional[str]:
        nonlocal created_tags, synthetic_tag_id, failed_tag_ensures
        clean_name = normalize_tag_name(tag_name)
        if not clean_name:
            return None

        key = normalize_text(clean_name)
        existing = tags_by_norm.get(key)
        if existing:
            return str(existing["id"])

        if csv_mode:
            nonlocal max_tag_id
            max_tag_id += 1
            new_tag = {"id": str(max_tag_id), "name": clean_name}
            if tags_has_category and category:
                new_tag["category"] = category
            tags_by_norm[key] = new_tag
            created_tags += 1
            merged_tags_rows.append(new_tag)
            return str(new_tag["id"])

        if not args.apply:
            synthetic = {"id": str(synthetic_tag_id), "name": clean_name}
            if tags_has_category and category:
                synthetic["category"] = category
            synthetic_tag_id -= 1
            tags_by_norm[key] = synthetic
            created_tags += 1
            return str(synthetic["id"])

        insert_payload = {"name": clean_name}
        if tags_has_category and category:
            insert_payload["category"] = category
        insert_error: Optional[Exception] = None
        try:
            assert supabase is not None
            response = supabase.table("tags").insert(insert_payload).execute()
            inserted = (response.data or [None])[0]
            if inserted and inserted.get("id") is not None:
                tags_by_norm[key] = inserted
                created_tags += 1
                return str(inserted["id"])
        except Exception as exc:
            insert_error = exc

        # Insert may return no representation or fail on duplicate; lookup by name robustly.
        assert supabase is not None
        inserted = find_tag_by_name(supabase, clean_name)
        if not inserted or inserted.get("id") is None:
            failed_tag_ensures += 1
            if insert_error is not None:
                print(f"[warn] Could not ensure tag '{clean_name}' ({insert_error})")
            else:
                print(f"[warn] Could not ensure tag '{clean_name}'")
            return None

        tags_by_norm[key] = inserted
        return str(inserted["id"])

    def match_place_id(row: DatasetRow) -> Tuple[Optional[str], str]:
        raw_id = row.values.get("id")
        if raw_id is not None:
            by_id = place_by_id.get(str(raw_id))
            if by_id:
                return str(by_id["id"]), "id"

        if row.gmaps_place_id:
            by_pid = place_by_gmaps_place_id.get(normalize_text(row.gmaps_place_id))
            if by_pid:
                return str(by_pid["id"]), "gmaps_place_id"

        if row.gmaps_uri:
            by_uri = place_by_gmaps_uri.get(normalize_text(row.gmaps_uri))
            if by_uri:
                return str(by_uri["id"]), "gmaps_uri"

        key = f"{normalize_text(row.place_name)}||{normalize_text(row.formatted_address)}"
        by_name_addr = place_by_name_address.get(key)
        if by_name_addr:
            return str(by_name_addr["id"]), "name+address"

        return None, "none"

    def process_row(row: DatasetRow, place_id: str, matched_by: str) -> None:
        nonlocal rows_with_no_tags, planned_links

        area_tag = infer_area_tag(row, geocoder)
        price_tag = infer_price_range_tag(row)
        cuisine_tags = infer_cuisine_tags(row)
        allergy_tags = infer_allergy_tags(row)

        proposed = [t for t in [area_tag, price_tag, *cuisine_tags, *allergy_tags] if t and is_english_tag(t)]
        # Preserve order, remove duplicates case-insensitively.
        seen_norm: Set[str] = set()
        deduped: List[str] = []
        for t in proposed:
            key = normalize_text(t)
            if key in seen_norm:
                continue
            seen_norm.add(key)
            deduped.append(t)

        if not deduped:
            rows_with_no_tags += 1
            report_rows.append(
                {
                    "row_num": row.row_num,
                    "place_name": row.place_name,
                    "matched": True,
                    "matched_by": matched_by,
                    "place_id": place_id,
                    "area_tag": area_tag or "",
                    "price_range_tag": price_tag,
                    "cuisine_tags": "|".join(cuisine_tags),
                    "allergy_tags": "|".join(allergy_tags),
                    "all_inferred_tags": "",
                    "new_links_planned_or_inserted": 0,
                }
            )
            return

        row_links = 0
        for tag_name in deduped:
            category = infer_tag_category(tag_name, area_tag)
            tag_id = ensure_tag_id(tag_name, category)
            if not tag_id:
                continue

            pair = (place_id, tag_id)
            if pair in existing_pairs:
                continue

            existing_pairs.add(pair)
            planned_links += 1
            row_links += 1
            new_link = {"place_id": place_id, "tag_id": tag_id}
            place_tags_to_insert.append(new_link)
            if csv_mode:
                merged_place_tags_rows.append(new_link)

        report_rows.append(
            {
                "row_num": row.row_num,
                "place_name": row.place_name,
                "matched": True,
                "matched_by": matched_by,
                "place_id": place_id,
                "area_tag": area_tag or "",
                "price_range_tag": price_tag,
                "cuisine_tags": "|".join(cuisine_tags),
                "allergy_tags": "|".join(allergy_tags),
                "all_inferred_tags": "|".join(deduped),
                "new_links_planned_or_inserted": row_links,
            }
        )

    if args.source == "dataset":
        for row in dataset_rows:
            place_id, matched_by = match_place_id(row)
            if not place_id:
                unmatched_rows += 1
                report_rows.append(
                    {
                        "row_num": row.row_num,
                        "place_name": row.place_name,
                        "matched": False,
                        "matched_by": matched_by,
                        "place_id": "",
                        "area_tag": "",
                        "cuisine_tags": "",
                        "allergy_tags": "",
                        "all_inferred_tags": "",
                        "new_links_planned_or_inserted": 0,
                    }
                )
                continue
            process_row(row, place_id, matched_by)
    else:
        for idx, place in enumerate(place_rows, start=1):
            place_id = str(place.get("id") or "")
            if not place_id:
                unmatched_rows += 1
                report_rows.append(
                    {
                        "row_num": idx,
                        "place_name": str(place.get("name") or ""),
                        "matched": False,
                        "matched_by": "places",
                        "place_id": "",
                        "area_tag": "",
                        "cuisine_tags": "",
                        "allergy_tags": "",
                        "all_inferred_tags": "",
                        "new_links_planned_or_inserted": 0,
                    }
                )
                continue
            row = PlaceRow(row_num=idx, values=place)
            process_row(row, place_id, "places")

    if not csv_mode and args.apply and place_tags_to_insert:
        assert supabase is not None
        for batch in chunked(place_tags_to_insert, 250):
            try:
                supabase.table("place_tags").insert(list(batch)).execute()
            except Exception:
                # Retry row-by-row so one bad row doesn't stop everything.
                for row in batch:
                    try:
                        supabase.table("place_tags").insert(row).execute()
                    except Exception:
                        failed_link_inserts += 1

    print("\nSummary")
    print(f"- Existing tags loaded: {len(tags_rows)}")
    print(f"- Existing place_tags loaded: {len(place_tag_rows)}")
    print(f"- New tags {'created' if (args.apply or csv_mode) else 'planned'}: {created_tags}")
    if args.apply and not csv_mode:
        print(f"- Failed tag ensure operations: {failed_tag_ensures}")
    print(f"- New place_tags {'inserted' if (args.apply or csv_mode) else 'planned'}: {planned_links}")
    if args.apply and not csv_mode:
        print(f"- Failed place_tags inserts: {failed_link_inserts}")
    unmatched_label = "Unmatched dataset rows" if args.source == "dataset" else "Rows missing place_id"
    print(f"- {unmatched_label}: {unmatched_rows}")
    print(f"- Rows with no inferred tags: {rows_with_no_tags}")

    if csv_mode:
        out_tags_csv = Path(args.out_tags_csv).expanduser() if args.out_tags_csv else Path.cwd() / "tags.generated.csv"
        out_place_tags_csv = (
            Path(args.out_place_tags_csv).expanduser() if args.out_place_tags_csv else Path.cwd() / "place_tags.generated.csv"
        )
        if not out_tags_csv.is_absolute():
            out_tags_csv = Path.cwd() / out_tags_csv
        if not out_place_tags_csv.is_absolute():
            out_place_tags_csv = Path.cwd() / out_place_tags_csv

        if not tags_fieldnames:
            tags_fieldnames = ["id", "name"]
        if not place_tags_fieldnames:
            place_tags_fieldnames = ["place_id", "tag_id"]

        if "id" in tags_fieldnames:
            merged_tags_rows = [r for r in merged_tags_rows if is_english_tag(str(r.get("name", "")))]
            merged_tags_rows.sort(key=lambda r: int(str(r.get("id"))))
            valid_tag_ids = {str(r.get("id")) for r in merged_tags_rows}
            merged_place_tags_rows = [r for r in merged_place_tags_rows if str(r.get("tag_id")) in valid_tag_ids]
        if "place_id" in place_tags_fieldnames and "tag_id" in place_tags_fieldnames:
            merged_place_tags_rows.sort(key=lambda r: (int(str(r.get("place_id"))), int(str(r.get("tag_id")))))

        write_csv_rows(out_tags_csv, tags_fieldnames, merged_tags_rows)
        write_csv_rows(out_place_tags_csv, place_tags_fieldnames, merged_place_tags_rows)
        print(f"- Updated tags CSV written: {out_tags_csv}")
        print(f"- Updated place_tags CSV written: {out_place_tags_csv}")

    if args.report:
        report_path = Path(args.report).expanduser()
        if not report_path.is_absolute():
            report_path = Path.cwd() / report_path
        report_path.parent.mkdir(parents=True, exist_ok=True)

        fieldnames = [
            "row_num",
            "place_name",
            "matched",
            "matched_by",
            "place_id",
            "area_tag",
            "price_range_tag",
            "cuisine_tags",
            "allergy_tags",
            "all_inferred_tags",
            "new_links_planned_or_inserted",
        ]
        if report_path.suffix.lower() == ".csv":
            with report_path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(report_rows)
        elif report_path.suffix.lower() == ".xlsx":
            try:
                from openpyxl import Workbook
            except Exception as exc:
                raise RuntimeError("openpyxl is required to write .xlsx reports") from exc

            wb = Workbook()
            ws = wb.active
            ws.title = "auto_tag_report"
            ws.append(fieldnames)
            for row in report_rows:
                ws.append([row.get(name, "") for name in fieldnames])
            wb.save(report_path)
        else:
            with report_path.open("w", encoding="utf-8") as f:
                json.dump(report_rows, f, ensure_ascii=False, indent=2)

        print(f"- Report written: {report_path}")


if __name__ == "__main__":
    main()
