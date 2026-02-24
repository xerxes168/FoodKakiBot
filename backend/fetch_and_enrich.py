"""
fetch_and_enrich.py
-------------------
Fetches the first 10 restaurants from Supabase, enriches each with Google Places API (New) data,
and exports a formatted Excel file.

Cuisine tagging strategy (in priority order):
  1. primaryType / types from Places API (New) — structured Google data e.g. "japanese_restaurant"
  2. editorialSummary — human-written blurb, often names cuisine explicitly
  3. reviews (top 5) — user text, keyword matched as last resort

Run:
    pip install -r requirements.txt
    python fetch_and_enrich.py                        # fetch 10 (default)
    python fetch_and_enrich.py --limit 50             # fetch 50
    python fetch_and_enrich.py --limit 0              # fetch all
    python fetch_and_enrich.py --limit 25 --output my_data.xlsx

Requires backend/.env with:
    GOOGLE_PLACES_API_KEY=<your key>
    SUPABASE_URL=https://...
    SUPABASE_ANON_KEY=sb_publishable_...
"""

import os
import re
import requests
from datetime import datetime
from supabase import create_client
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_ANON_KEY"]
GOOGLE_API_KEY = os.environ["GOOGLE_PLACES_API_KEY"]

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

PRICE_MAP = {
    0: "Free",
    1: "$ (Budget <$15)",
    2: "$$ (Moderate $15-$30)",
    3: "$$$ (Expensive $30-$60)",
    4: "$$$$ (Very Expensive $60+)",
}

# ── Places API (New) primaryType → human-readable cuisine label ───────────────
# Full reference: https://developers.google.com/maps/documentation/places/web-service/place-types
PRIMARY_TYPE_MAP = {
    # Specific cuisine types (New API)
    "afghani_restaurant": "Afghan",
    "african_restaurant": "African",
    "american_restaurant": "American",
    "asian_restaurant": "Asian",
    "australian_restaurant": "Australian",
    "austrian_restaurant": "Austrian",
    "bangladeshi_restaurant": "Bangladeshi",
    "barbecue_restaurant": "BBQ",
    "belgian_restaurant": "Belgian",
    "brazilian_restaurant": "Brazilian",
    "british_restaurant": "British",
    "brunch_restaurant": "Brunch",
    "buffet_restaurant": "Buffet",
    "cafe": "Cafe",
    "cambodian_restaurant": "Cambodian",
    "caribbean_restaurant": "Caribbean",
    "chinese_restaurant": "Chinese",
    "comfort_food_restaurant": "Comfort Food",
    "contemporary_restaurant": "Contemporary",
    "deli": "Deli",
    "dim_sum_restaurant": "Dim Sum",
    "diner": "Diner",
    "ethiopian_restaurant": "Ethiopian",
    "fast_food_restaurant": "Fast Food",
    "french_restaurant": "French",
    "fusion_restaurant": "Fusion",
    "german_restaurant": "German",
    "greek_restaurant": "Greek",
    "hamburger_restaurant": "Burgers",
    "hawaiian_restaurant": "Hawaiian",
    "hokkien_restaurant": "Hokkien",
    "hot_pot_restaurant": "Hotpot / Steamboat",
    "indian_restaurant": "Indian",
    "indonesian_restaurant": "Indonesian",
    "international_restaurant": "International",
    "italian_restaurant": "Italian",
    "japanese_restaurant": "Japanese",
    "korean_restaurant": "Korean",
    "latin_american_restaurant": "Latin American",
    "lebanese_restaurant": "Lebanese",
    "lunch_restaurant": "Lunch",
    "malay_restaurant": "Malay",
    "mediterranean_restaurant": "Mediterranean",
    "mexican_restaurant": "Mexican",
    "middle_eastern_restaurant": "Middle Eastern",
    "modern_european_restaurant": "Modern European",
    "moroccan_restaurant": "Moroccan",
    "noodle_restaurant": "Noodles",
    "pakistani_restaurant": "Pakistani",
    "pan_asian_restaurant": "Pan-Asian",
    "pizza_restaurant": "Pizza",
    "poke_restaurant": "Poke",
    "portuguese_restaurant": "Portuguese",
    "ramen_restaurant": "Ramen",
    "sandwich_shop": "Sandwiches",
    "seafood_restaurant": "Seafood",
    "singaporean_restaurant": "Singaporean",
    "soup_restaurant": "Soup",
    "spanish_restaurant": "Spanish",
    "sri_lankan_restaurant": "Sri Lankan",
    "steak_house": "Steakhouse",
    "sushi_restaurant": "Sushi",
    "taiwanese_restaurant": "Taiwanese",
    "tapas_restaurant": "Tapas",
    "tea_house": "Tea House",
    "tex_mex_restaurant": "Tex-Mex",
    "thai_restaurant": "Thai",
    "turkish_restaurant": "Turkish",
    "vegan_restaurant": "Vegan",
    "vegetarian_restaurant": "Vegetarian",
    "vietnamese_restaurant": "Vietnamese",
    "wine_bar": "Wine Bar",
    # Generic fallbacks still worth capturing
    "restaurant": "Restaurant",
    "food": "Food",
    "bakery": "Bakery",
    "bar": "Bar",
    "ice_cream_shop": "Ice Cream",
    "bubble_tea_store": "Bubble Tea",
    "dessert_shop": "Dessert",
    "juice_shop": "Juice",
    "meal_takeaway": "Takeaway",
    "meal_delivery": "Delivery",
}

# Keywords for fallback inference from editorial_summary + reviews text
CUISINE_KEYWORDS = {
    "Japanese":     ["japanese","sushi","sashimi","ramen","udon","soba","yakitori","izakaya","tempura","tonkatsu","donburi","omakase","katsu","wagyu","gyoza","onigiri"],
    "Korean":       ["korean","kimchi","bibimbap","bulgogi","tteokbokki","kbbq","korean bbq","banchan","samgyeopsal","jjajangmyeon","doenjang"],
    "Chinese":      ["chinese","dim sum","dumpling","wonton","char siew","cantonese","szechuan","sichuan","hotpot","steamboat","zi char","xiao long bao","claypot","peking","wok"],
    "Indian":       ["indian","biryani","tandoori","naan","curry","masala","dosa","prata","thosai","tikka","paneer","briyani","chaat"],
    "Malay":        ["malay","nasi lemak","satay","rendang","mee rebus","nasi padang","padang","laksa","otah","mee goreng"],
    "Thai":         ["thai","tom yum","pad thai","green curry","som tam","basil","thai basil","khao pad"],
    "Vietnamese":   ["vietnamese","pho","banh mi","bun cha","spring roll","com tam","boba","viet"],
    "Western":      ["western","steak","burger","grill","bistro","barbecue","ribs","pub","brunch","roast","chips","fish and chips"],
    "Italian":      ["italian","pasta","pizza","risotto","carbonara","lasagna","trattoria","gelato","tiramisu","osso buco","arancini"],
    "Mexican":      ["mexican","taco","burrito","quesadilla","nachos","guacamole","enchilada","fajita"],
    "Seafood":      ["seafood","crab","lobster","oyster","prawn","clam","fish","mussel","scallop","chilli crab"],
    "Vegetarian":   ["vegetarian","vegan","plant-based","meat-free","tofu","tempeh"],
    "Halal":        ["halal","muslim-friendly","no pork no lard","nasi padang","mamak","murtabak","kebab","shawarma"],
    "Dessert":      ["dessert","cake","gelato","ice cream","pastry","brownie","tart","pudding","crepe","waffle","churros"],
    "Cafe":         ["cafe","coffee","latte","espresso","flat white","cappuccino","brunch","barista"],
    "Mala":         ["mala","ma la","spicy pot","xiang guo","sichuan spicy","dry pot"],
    "Bubble Tea":   ["bubble tea","boba","milk tea","gong cha","koi","liho","playmade","chagee","heytea","taro"],
    "Fast Food":    ["fast food","mcdonald","kfc","subway","burger king","jollibee","popeyes"],
    "Singaporean":  ["hawker","kopitiam","chicken rice","char kway teow","bak kut teh","laksa","rojak","carrot cake","wanton mee"],
}

HALAL_KEYWORDS = [
    "halal","muslim-friendly","no pork no lard","nasi padang","mamak",
    "murtabak","kebab","shawarma","warong","warung","muslim owned"
]


# ── Google Types → flat tags (non-cuisine attributes) ────────────────────────
# These complement cuisine tags — capturing dining format, amenities, etc.
GOOGLE_TYPE_TO_TAG = {
    # Dining format
    "meal_takeaway":        "Takeaway",
    "meal_delivery":        "Delivery",
    "fast_food_restaurant": "Fast Food",
    "buffet_restaurant":    "Buffet",
    "brunch_restaurant":    "Brunch",
    "cafe":                 "Cafe",
    "bar":                  "Bar",
    "wine_bar":             "Wine Bar",
    "bakery":               "Bakery",
    "dessert_shop":         "Dessert",
    "ice_cream_shop":       "Ice Cream",
    "bubble_tea_store":     "Bubble Tea",
    "juice_shop":           "Juice Bar",
    "tea_house":            "Tea House",
    "diner":                "Diner",
    "food_court":           "Food Court",
    "hawker_center":        "Hawker Centre",
    # Venue type
    "night_club":           "Nightclub",
    "tourist_attraction":   "Tourist Spot",
    "shopping_mall":        "In Mall",
    # Dietary
    "vegan_restaurant":     "Vegan",
    "vegetarian_restaurant":"Vegetarian",
}


def tags_from_google_types(types: list) -> list[str]:
    """Extract non-cuisine descriptive tags from the full types[] array."""
    seen = set()
    tags = []
    for t in types:
        label = GOOGLE_TYPE_TO_TAG.get(t)
        if label and label not in seen:
            seen.add(label)
            tags.append(label)
    return tags


def price_from_range(price_range: dict) -> tuple[int | None, str]:
    """
    Parse priceRange (Places API New) when priceLevel is missing.
    priceRange = {
        "startPrice": {"currencyCode": "SGD", "units": "15"},
        "endPrice":   {"currencyCode": "SGD", "units": "40"}
    }
    Returns (price_num, price_label) using midpoint heuristic in SGD.
    """
    if not price_range:
        return None, "N/A"
    try:
        start = float((price_range.get("startPrice") or {}).get("units", 0) or 0)
        end   = float((price_range.get("endPrice")   or {}).get("units", 0) or 0)
        mid   = (start + end) / 2 if end else start
        raw_label = f"~SGD ${int(start)}–${int(end)}" if end else f"~SGD ${int(start)}+"
        if mid <= 15:
            return 1, f"$ (Budget <$15) [{raw_label}]"
        elif mid <= 30:
            return 2, f"$$ (Moderate $15–$30) [{raw_label}]"
        elif mid <= 60:
            return 3, f"$$$ (Expensive $30–$60) [{raw_label}]"
        else:
            return 4, f"$$$$ (Very Expensive $60+) [{raw_label}]"
    except (TypeError, ValueError):
        return None, "N/A"


# ── Google Places API (New) ───────────────────────────────────────────────────

PLACES_NEW_BASE = "https://places.googleapis.com/v1/places"

NEW_API_FIELDS = ",".join([
    "id",
    "displayName",
    "formattedAddress",
    "location",
    "rating",
    "userRatingCount",
    "priceLevel",
    "priceRange",
    "primaryType",
    "primaryTypeDisplayName",
    "types",
    "editorialSummary",
    "regularOpeningHours",
    "internationalPhoneNumber",
    "websiteUri",
    "googleMapsUri",
    "photos",
    "businessStatus",
    "servesBeer",
    "servesBreakfast",
    "servesBrunch",
    "servesDinner",
    "servesLunch",
    "servesVegetarianFood",
    "accessibilityOptions",
    "reviews",
    "dineIn",
    "takeout",
    "delivery",
    "reservable",
    "goodForChildren",
    "goodForGroups",
    "liveMusic",
    "outdoorSeating",
])


def get_place_details_new(place_id: str) -> dict:
    """Call Places API (New) — returns richer structured data incl. primaryType."""
    url = f"{PLACES_NEW_BASE}/{place_id}"
    headers = {
        "X-Goog-Api-Key": GOOGLE_API_KEY,
        "X-Goog-FieldMask": NEW_API_FIELDS,
    }
    r = requests.get(url, headers=headers, timeout=20)
    if r.status_code != 200:
        print(f"  WARNING: Places API (New) error {r.status_code} for {place_id}: {r.text[:200]}")
        return {}
    return r.json()


# ── Cuisine resolution ────────────────────────────────────────────────────────

def cuisine_from_primary_type(primary_type: str) -> str | None:
    """Return a clean cuisine label from primaryType if it's specific enough."""
    if not primary_type:
        return None
    label = PRIMARY_TYPE_MAP.get(primary_type)
    # Skip generic non-informative types
    if label in ("Restaurant", "Food", "Takeaway", "Delivery", None):
        return None
    return label


def cuisine_from_text(text: str) -> list[str]:
    """Keyword-based fallback — scans editorial summary + review text."""
    t = text.lower()
    return [cuisine for cuisine, kws in CUISINE_KEYWORDS.items() if any(kw in t for kw in kws)]


def resolve_cuisine(details: dict) -> tuple[str, str]:
    """
    Returns (cuisine_tags, cuisine_source) where source is one of:
      'primaryType' | 'types' | 'editorial' | 'reviews' | 'inferred'
    """
    primary_type = details.get("primaryType", "")
    all_types = details.get("types", [])

    # 1) Try primaryType first
    label = cuisine_from_primary_type(primary_type)
    if label:
        return label, "primaryType"

    # 2) Scan all types[] for specific cuisine entries
    type_labels = [
        PRIMARY_TYPE_MAP[t] for t in all_types
        if t in PRIMARY_TYPE_MAP and PRIMARY_TYPE_MAP[t] not in ("Restaurant", "Food", "Takeaway", "Delivery")
    ]
    if type_labels:
        return ", ".join(list(dict.fromkeys(type_labels))), "types[]"

    # 3) Editorial summary
    editorial = (details.get("editorialSummary") or {}).get("text", "")
    if editorial:
        matched = cuisine_from_text(editorial)
        if matched:
            return ", ".join(matched[:3]), "editorial"

    # 4) Reviews text
    reviews = details.get("reviews") or []
    review_text = " ".join(
        (r.get("text") or {}).get("text", "") for r in reviews[:5]
    )
    if review_text:
        matched = cuisine_from_text(review_text)
        if matched:
            return ", ".join(list(dict.fromkeys(matched))[:3]), "reviews"

    # 5) Last resort — display name of primaryType even if generic
    display = (details.get("primaryTypeDisplayName") or {}).get("text", "")
    if display:
        return display, "primaryTypeDisplayName"

    return "Unknown", "none"


def infer_halal(name: str, types: list, editorial: str, reviews_text: str) -> str:
    text = " ".join([name or "", " ".join(types or []), editorial or "", reviews_text or ""]).lower()
    return "Yes" if any(kw in text for kw in HALAL_KEYWORDS) else "Unknown"


# ── Supabase fetch ────────────────────────────────────────────────────────────

def fetch_restaurants(limit=10):
    q = supabase.table("places").select(
        "id,gmaps_place_id,name,address,latitude,longitude,"
        "rating,user_rating_count,price_level,types,"
        "editorial_summary,gmaps_uri,website_uri"
    )
    if limit is not None:
        q = q.limit(limit)
    return q.execute().data


# ── Row builder ───────────────────────────────────────────────────────────────

PRICE_LEVEL_MAP_NEW = {
    "PRICE_LEVEL_FREE": (0, "Free"),
    "PRICE_LEVEL_INEXPENSIVE": (1, "$ (Budget <$15)"),
    "PRICE_LEVEL_MODERATE": (2, "$$ (Moderate $15-$30)"),
    "PRICE_LEVEL_EXPENSIVE": (3, "$$$ (Expensive $30-$60)"),
    "PRICE_LEVEL_VERY_EXPENSIVE": (4, "$$$$ (Very Expensive $60+)"),
}


def build_row(r: dict, details: dict) -> dict:
    name = (details.get("displayName") or {}).get("text") or r.get("name", "")
    types = details.get("types") or r.get("types") or []
    editorial = (details.get("editorialSummary") or {}).get("text") or r.get("editorial_summary") or ""

    # Reviews text for halal + cuisine fallback
    reviews = details.get("reviews") or []
    reviews_text = " ".join((rv.get("text") or {}).get("text", "") for rv in reviews[:5])

    # Price — prefer priceLevel enum, fall back to priceRange, then DB value
    price_raw = details.get("priceLevel") or ""
    if price_raw in PRICE_LEVEL_MAP_NEW:
        price_num, price_label = PRICE_LEVEL_MAP_NEW[price_raw]
        price_source = "priceLevel"
    else:
        # Try priceRange (shown on Google website when priceLevel is absent)
        price_num, price_label = price_from_range(details.get("priceRange"))
        if price_num is not None:
            price_source = "priceRange"
        else:
            price_num = r.get("price_level")
            price_label = PRICE_MAP.get(price_num, "N/A") if price_num is not None else "N/A"
            price_source = "database" if price_num is not None else "none"

    rating = details.get("rating") or r.get("rating")
    rating_count = details.get("userRatingCount") or r.get("user_rating_count")

    # Opening hours
    oh = details.get("regularOpeningHours") or {}
    open_now = ("Yes" if oh.get("openNow") else "No") if "openNow" in oh else "N/A"
    weekday_hours = "\n".join(oh.get("weekdayDescriptions", [])) if oh.get("weekdayDescriptions") else "N/A"

    # Meal types
    meals = [m for m, flag in [
        ("Breakfast", details.get("servesBreakfast")),
        ("Brunch", details.get("servesBrunch")),
        ("Lunch", details.get("servesLunch")),
        ("Dinner", details.get("servesDinner")),
    ] if flag]

    # Cuisine
    cuisine_tags, cuisine_source = resolve_cuisine(details)

    # Tags from Google Types (dining format, venue attributes)
    google_type_tags = tags_from_google_types(types)

    # Halal
    halal = infer_halal(name, types, editorial, reviews_text)

    # Dietary
    dietary = []
    if details.get("servesVegetarianFood"):
        dietary.append("Vegetarian-Friendly")
    if halal == "Yes":
        dietary.append("Halal")

    # Accessibility
    accessibility = details.get("accessibilityOptions") or {}
    wheelchair = "Yes" if accessibility.get("wheelchairAccessibleEntrance") else "N/A"

    # Amenity tags from extra API fields
    amenity_tags = []
    if details.get("dineIn"):       amenity_tags.append("Dine-In")
    if details.get("takeout"):      amenity_tags.append("Takeaway")
    if details.get("delivery"):     amenity_tags.append("Delivery")
    if details.get("reservable"):   amenity_tags.append("Reservable")
    if details.get("outdoorSeating"): amenity_tags.append("Outdoor Seating")
    if details.get("liveMusic"):    amenity_tags.append("Live Music")
    if details.get("goodForGroups"): amenity_tags.append("Good for Groups")
    if details.get("goodForChildren"): amenity_tags.append("Family-Friendly")

    # Combine all tags (deduplicated)
    all_tags = list(dict.fromkeys(
        [cuisine_tags] + google_type_tags + amenity_tags
    )) if cuisine_tags != "Unknown" else list(dict.fromkeys(google_type_tags + amenity_tags))

    # Photos
    photos = details.get("photos") or []

    return {
        "DB ID": r.get("id"),
        "Name": name,
        "Address": details.get("formattedAddress") or r.get("address") or "",
        "Latitude": r.get("latitude"),
        "Longitude": r.get("longitude"),
        "Business Status": details.get("businessStatus") or "N/A",
        "Rating (Google)": rating,
        "Total Ratings": rating_count,
        "Price Level (1-4)": price_num,
        "Price Category": price_label,
        "Price Source": price_source,
        "Cuisine Tags": cuisine_tags,
        "Cuisine Source": cuisine_source,
        "Google Type Tags": ", ".join(google_type_tags) if google_type_tags else "N/A",
        "Amenity Tags": ", ".join(amenity_tags) if amenity_tags else "N/A",
        "All Tags": ", ".join(all_tags) if all_tags else "N/A",
        "Primary Type (Raw)": details.get("primaryType") or "",
        "Primary Type (Display)": (details.get("primaryTypeDisplayName") or {}).get("text") or "",
        "Halal": halal,
        "Meal Types Served": ", ".join(meals) if meals else "N/A",
        "Dietary Info": ", ".join(dietary) if dietary else "Not Specified",
        "Serves Alcohol": "Yes" if details.get("servesBeer") else ("No" if details else "N/A"),
        "Wheelchair Accessible": wheelchair,
        "Open Now": open_now,
        "Opening Hours": weekday_hours,
        "Phone": details.get("internationalPhoneNumber") or "",
        "Google Maps URL": details.get("googleMapsUri") or r.get("gmaps_uri") or "",
        "Website": details.get("websiteUri") or r.get("website_uri") or "",
        "Editorial Summary": editorial,
        "All Google Types": ", ".join(types),
        "Photos Available": len(photos),
        "Google Place ID": r.get("gmaps_place_id") or "",
    }


# ── Excel generation ──────────────────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

COL_WIDTHS = {
    "DB ID": 8, "Name": 30, "Address": 40, "Latitude": 12, "Longitude": 12,
    "Business Status": 18, "Rating (Google)": 15, "Total Ratings": 14,
    "Price Level (1-4)": 16, "Price Category": 30, "Price Source": 14,
    "Cuisine Tags": 35, "Cuisine Source": 20,
    "Google Type Tags": 35, "Amenity Tags": 40, "All Tags": 50,
    "Primary Type (Raw)": 28, "Primary Type (Display)": 28,
    "Halal": 10, "Meal Types Served": 28, "Dietary Info": 28,
    "Serves Alcohol": 14, "Wheelchair Accessible": 20,
    "Open Now": 10, "Opening Hours": 55, "Phone": 22,
    "Google Maps URL": 45, "Website": 40, "Editorial Summary": 55,
    "All Google Types": 50, "Photos Available": 16, "Google Place ID": 32,
}

# Highlight "Cuisine Source" column to make provenance visible
SOURCE_COLORS = {
    "primaryType": "C8E6C9",       # green  — most reliable
    "types[]": "DCEDC8",           # light green
    "editorial": "FFF9C4",         # yellow — inferred
    "reviews": "FFE0B2",           # orange — weakest structured
    "primaryTypeDisplayName": "F3E5F5",  # purple
    "inferred": "FFCCBC",
    "none": "FFCDD2",              # red    — unknown
}


def create_excel(rows: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Restaurant Data"

    headers = list(rows[0].keys())
    cuisine_source_col = headers.index("Cuisine Source") + 1

    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color="E85D04")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="medium", color="BF360C"),
            right=Side(style="medium", color="BF360C"),
            top=Side(style="medium", color="BF360C"),
            bottom=Side(style="medium", color="BF360C"),
        )

    ws.row_dimensions[1].height = 40

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        alt = row_idx % 2 == 0
        ws.row_dimensions[row_idx].height = 60
        source = row.get("Cuisine Source", "none")

        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.font = Font(name="Arial", size=10)

            if col_idx == cuisine_source_col:
                color = SOURCE_COLORS.get(source, "FFFFFF")
                cell.fill = PatternFill("solid", start_color=color)
                cell.font = Font(name="Arial", size=10, bold=True)
            elif alt:
                cell.fill = PatternFill("solid", start_color="FFF3E0")

    # Column widths
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(h, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 42

    ws2["A1"] = "FoodKakiBot — Restaurant Enrichment Report"
    ws2["A1"].font = Font(name="Arial", bold=True, size=16, color="E85D04")
    ws2.merge_cells("A1:B1")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 30

    summary_items = [
        ("Restaurants Fetched", len(rows)),
        ("Average Rating", f"=AVERAGE('Restaurant Data'!G2:G{len(rows)+1})"),
        ("Data Source", "Supabase PostgreSQL + Google Places API (New)"),
        ("Generated On", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for r_idx, (label, value) in enumerate(summary_items, 3):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(name="Arial", bold=True, size=11)
        ws2.cell(row=r_idx, column=2, value=value).font = Font(name="Arial", size=11)

    # Cuisine source breakdown
    ws2["A8"] = "Cuisine Source Breakdown"
    ws2["A8"].font = Font(name="Arial", bold=True, size=12, color="E85D04")
    ws2["B8"] = "Count"
    ws2["B8"].font = Font(name="Arial", bold=True, size=12, color="E85D04")

    source_counts: dict = {}
    for row in rows:
        s = row.get("Cuisine Source", "none")
        source_counts[s] = source_counts.get(s, 0) + 1

    for r_idx, (src, cnt) in enumerate(sorted(source_counts.items(), key=lambda x: -x[1]), 9):
        lc = ws2.cell(row=r_idx, column=1, value=src)
        lc.font = Font(name="Arial", size=11)
        color = SOURCE_COLORS.get(src, "FFFFFF")
        lc.fill = PatternFill("solid", start_color=color)
        ws2.cell(row=r_idx, column=2, value=cnt).font = Font(name="Arial", size=11)

    # Cuisine tag breakdown
    next_row = 9 + len(source_counts) + 2
    ws2.cell(row=next_row, column=1, value="Cuisine Tag Breakdown").font = Font(name="Arial", bold=True, size=12, color="E85D04")
    ws2.cell(row=next_row, column=2, value="Count").font = Font(name="Arial", bold=True, size=12, color="E85D04")

    cuisine_counts: dict = {}
    for row in rows:
        for c in row.get("Cuisine Tags", "").split(", "):
            c = c.strip()
            if c and c != "Unknown":
                cuisine_counts[c] = cuisine_counts.get(c, 0) + 1

    for r_idx, (cuisine, count) in enumerate(sorted(cuisine_counts.items(), key=lambda x: -x[1]), next_row + 1):
        ws2.cell(row=r_idx, column=1, value=cuisine).font = Font(name="Arial", size=11)
        ws2.cell(row=r_idx, column=2, value=count).font = Font(name="Arial", size=11)

    # Legend
    legend_row = next_row + len(cuisine_counts) + 3
    ws2.cell(row=legend_row, column=1, value="Cuisine Source Legend").font = Font(name="Arial", bold=True, size=12, color="E85D04")
    legend = [
        ("primaryType", "Google's structured primary type — most reliable"),
        ("types[]", "From Google's types array — reliable"),
        ("editorial", "Inferred from editorial summary text"),
        ("reviews", "Inferred from user reviews text"),
        ("primaryTypeDisplayName", "Google display name fallback"),
        ("none", "Could not determine cuisine"),
    ]
    for r_idx, (src, desc) in enumerate(legend, legend_row + 1):
        lc = ws2.cell(row=r_idx, column=1, value=src)
        lc.font = Font(name="Arial", size=10, bold=True)
        lc.fill = PatternFill("solid", start_color=SOURCE_COLORS.get(src, "FFFFFF"))
        ws2.cell(row=r_idx, column=2, value=desc).font = Font(name="Arial", size=10)

    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    print(f"Saved: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="FoodKakiBot - Restaurant Data Enrichment")
    parser.add_argument(
        "--limit", type=int, default=10,
        help="Number of restaurants to fetch from Supabase (default: 10, use 0 for all)"
    )
    parser.add_argument(
        "--output", type=str, default="restaurant_enriched_data.xlsx",
        help="Output Excel filename (default: restaurant_enriched_data.xlsx)"
    )
    args = parser.parse_args()

    limit = args.limit if args.limit > 0 else None

    print("FoodKakiBot - Restaurant Data Enrichment (Places API New)")
    print("=" * 60)
    print(f"Fetching {'all' if limit is None else limit} restaurants from Supabase...")
    restaurants = fetch_restaurants(limit)
    print(f"Fetched {len(restaurants)} restaurants\n")

    enriched_rows = []
    for r in restaurants:
        pid = r.get("gmaps_place_id", "")
        name = r.get("name", "Unknown")
        print(f"  Fetching: {name}")
        details = get_place_details_new(pid) if pid else {}
        if details:
            cuisine, source = resolve_cuisine(details)
            print(f"    Cuisine: {cuisine} (source: {source})")
        row = build_row(r, details)
        enriched_rows.append(row)

    print(f"\nBuilding Excel file...")
    create_excel(enriched_rows, args.output)
    print(f"Done! -> {args.output}")


if __name__ == "__main__":
    main()
